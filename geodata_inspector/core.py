"""
DuckDB-optimized Geodata Inspector
==================================
This module provides DuckDB-accelerated versions of the geodata inspection functions.
DuckDB offers significant performance improvements for:
- CSV reading (10-100x faster than pandas for large files)
- SQL-based aggregations and analysis
- Spatial operations via the spatial extension

Usage:
    python core.py [file_path]
"""

import os
import time
import duckdb
import pandas as pd
import geopandas as gpd
import numpy as np
from datetime import datetime
from shapely.geometry import Point, LineString, shape
from shapely import wkt
from shapely.validation import make_valid
import re
import json

# Import spatial metrics (reuse existing)
from . import spatial 
from .spatial import taux_de_remplissage, complexite_moyenne, pourcentage_geometries_dupliquees

# ============================================================================
# CONFIGURATION & STATE
# ============================================================================
summary_rows = []
last_gdf = None
last_dt_raw = None  # Fine-grained temporal data for re-computation with filters

# Initialize DuckDB with spatial extension
def get_duckdb_connection():
    """Create DuckDB connection with spatial extension loaded."""
    conn = duckdb.connect(':memory:')
    conn.execute("INSTALL spatial; LOAD spatial;")
    return conn


# ============================================================================
# DUCKDB UTILITY FUNCTIONS
# ============================================================================
def get_file_metadata(filepath):
    """Extract file metadata."""
    stats = os.stat(filepath)
    return {
        "Nom du fichier": os.path.basename(filepath),
        "Taille (Ko)": round(stats.st_size / 1024, 2),
        "Date de création du fichier (Y-M-D)": datetime.fromtimestamp(stats.st_ctime).strftime('%Y-%m-%d')
    }


def detect_csv_dialect_duckdb(filepath, sample_size=10000):
    """
    Detect CSV encoding and delimiter using DuckDB's sniffer.
    DuckDB automatically handles most CSV formats.
    """
    conn = duckdb.connect(':memory:')

    # Try common delimiters
    for delimiter in [',', ';', '\t', '|']:
        try:
            # Use DuckDB's read_csv with explicit delimiter
            result = conn.execute(f"""
                SELECT * FROM read_csv('{filepath}',
                    delim='{delimiter}',
                    header=true,
                    sample_size={sample_size},
                    ignore_errors=true
                ) LIMIT 5
            """).fetchdf()

            # If we got more than 1 column, likely correct delimiter
            if len(result.columns) > 1:
                conn.close()
                return {'delimiter': delimiter, 'encoding': 'utf-8'}
        except Exception:
            continue

    conn.close()
    return {'delimiter': ',', 'encoding': 'utf-8'}


def completeness_score_duckdb(conn, table_name):
    """
    Calculate completeness score using DuckDB SQL.
    Much faster than pandas for large datasets.
    """
    # Get column names
    columns = conn.execute(f"DESCRIBE {table_name}").fetchdf()['column_name'].tolist()

    if not columns:
        return {"Score de complétude moyen": None, "Score de complétude std": None}

    # Build SQL for null counts per column
    null_counts_sql = ", ".join([
        f"SUM(CASE WHEN \"{col}\" IS NULL THEN 1 ELSE 0 END)::DOUBLE / COUNT(*)::DOUBLE as null_ratio_{i}"
        for i, col in enumerate(columns)
    ])

    result = conn.execute(f"SELECT {null_counts_sql} FROM {table_name}").fetchone()

    if result is None:
        return {"Score de complétude moyen": None, "Score de complétude std": None}

    null_ratios = [r for r in result if r is not None]

    if not null_ratios:
        return {"Score de complétude moyen": 1.0, "Score de complétude std": 0}

    mean_completeness = 1 - np.mean(null_ratios)
    std_completeness = np.std(null_ratios)

    return {
        "_table": True,
        "data": [
            {
                "Score de complétude moyen (%)": round(mean_completeness * 100, 1),
                "Score de complétude std (%)":   round(std_completeness * 100, 1),
            }
        ]
    }

def completeness_score_duckdb_cols(conn, table_name, columns):
    """Calculate completeness score restricted to a specific list of columns."""
    if not columns:
        return {"_table": True, "data": [{"Score de complétude moyen (%)": None, "Score de complétude std (%)": None}]}

    null_counts_sql = ", ".join([
        f"SUM(CASE WHEN \"{col}\" IS NULL THEN 1 ELSE 0 END)::DOUBLE / COUNT(*)::DOUBLE as null_ratio_{i}"
        for i, col in enumerate(columns)
    ])

    result = conn.execute(f"SELECT {null_counts_sql} FROM {table_name}").fetchone()

    if result is None:
        return {"Score de complétude moyen": None, "Score de complétude std": None}

    null_ratios = [r for r in result if r is not None]
    mean_completeness = 1 - np.mean(null_ratios)
    std_completeness = np.std(null_ratios)

    return {
        "_table": True,
        "data": [
            {
                "Score de complétude moyen (%)": round(mean_completeness * 100, 1),
                "Score de complétude std (%)":   round(std_completeness * 100, 1),
            }
        ]
    }

def build_columns_detail_duckdb(conn, table_name, limit=5):
    """Build column details using DuckDB."""
    schema = conn.execute(f"DESCRIBE {table_name}").fetchdf()

    details = []
    for _, row in schema.iterrows():
        col_name = row['column_name']
        col_type = row['column_type']

        # Get sample value and null count
        try:
            # For geometry columns, convert WKB to readable WKT
            if col_type.upper() in ('GEOMETRY', 'BLOB') or 'GEOM' in col_name.upper():
                sample = conn.execute(f"""
                    SELECT ST_AsText("{col_name}") FROM {table_name}
                    WHERE "{col_name}" IS NOT NULL
                    LIMIT 1
                """).fetchone()
            else:
                sample = conn.execute(f"""
                    SELECT "{col_name}" FROM {table_name}
                    WHERE "{col_name}" IS NOT NULL
                    LIMIT 1
                """).fetchone()
            sample_val = str(sample[0]) if sample else "N/A"
            
            null_count = conn.execute(f"""
                SELECT COUNT(*) FROM {table_name} WHERE "{col_name}" IS NULL
            """).fetchone()[0]
        except Exception:
            sample_val = "N/A"
            null_count = 0

        details.append({
            "Colonne": col_name,
            "Exemple": sample_val[:100] if len(sample_val) > 100 else sample_val,
            "Type": col_type,
            "Valeurs manquantes": null_count
        })

    return details


# ============================================================================
# GEOGRAPHIC DETECTION (optimized for any listed zones in config.py file)
# ============================================================================

def detect_geo_join_keys_duckdb(conn, table_name, geo_key_patterns=None):
    """
    Detect geographic join key columns using DuckDB.
    Uses config-driven patterns if provided, falls back to
    length-based heuristics for unknown localisations.
    """
    import re
    candidates = []
    schema = conn.execute(f"DESCRIBE {table_name}").fetchdf()

    # ── CONFIG-DRIVEN DETECTION (when patterns are provided) ────────────
    if geo_key_patterns:
        for _, row in schema.iterrows():
            col = row['column_name']
            col_lc = col.lower()

            for key_def in geo_key_patterns:
                # Check if any pattern substring matches the column name
                name_match = any(p in col_lc for p in key_def["patterns"])
                if not name_match:
                    continue

                try:
                    # Skip all-null columns
                    null_check = conn.execute(f"""
                        SELECT COUNT(*) - COUNT("{col}") as null_count,
                               COUNT(*) as total
                        FROM {table_name}
                    """).fetchone()
                    if null_check[0] == null_check[1]:
                        continue

                    # Value format validation if defined
                    fmt = key_def.get("value_format")
                    if fmt:
                        # Sample up to 100 non-null values and check match rate
                        samples = conn.execute(f"""
                            SELECT CAST("{col}" AS VARCHAR)
                            FROM {table_name}
                            WHERE "{col}" IS NOT NULL
                            LIMIT 100
                        """).fetchdf().iloc[:, 0].tolist()

                        if samples:
                            match_rate = sum(
                                1 for v in samples
                                if re.match(fmt, str(v).strip())
                            ) / len(samples)
                            if match_rate < 0.5:
                                continue  # values don't look like this key type

                    candidates.append({"col": col, "label": key_def['label']})

                except Exception:
                    continue

            # Avoid duplicate entries for the same column
            seen_cols = set()
            deduped = []
            for c in candidates:
                col_part = c["col"]
                if col_part not in seen_cols:
                    seen_cols.add(col_part)
                    deduped.append(c)
            candidates = deduped

        return candidates

    # ── FALLBACK: LENGTH-BASED HEURISTICS (original logic) ──────────────
    pattern = r'(dep|reg|insee|com|code|postal|commune|departement|département)'
    for _, row in schema.iterrows():
        col = row['column_name']
        col_lc = col.lower()
        match = re.search(pattern, col_lc)
        if not match:
            continue
        try:
            stats = conn.execute(f"""
                SELECT
                    AVG(LENGTH(CAST("{col}" AS VARCHAR))) as avg_len,
                    COUNT(*) - COUNT("{col}") as null_count,
                    COUNT(*) as total
                FROM {table_name}
            """).fetchone()
            if stats[2] == stats[1]:
                continue
            avg_len = stats[0] or 0
            is_numeric = 'INT' in row['column_type'].upper() or 'DOUBLE' in row['column_type'].upper()
            w_match = col_lc[match.start():match.end()]
            if is_numeric:
                clean_len = conn.execute(f"""
                    SELECT AVG(LENGTH(CAST(FLOOR(CAST("{col}" AS DOUBLE)) AS BIGINT)::VARCHAR))
                    FROM {table_name}
                    WHERE "{col}" IS NOT NULL
                """).fetchone()[0] or 0
            else:
                clean_len = avg_len
            code_in_parens_len = conn.execute(f"""
                SELECT AVG(LENGTH(REGEXP_EXTRACT(CAST("{col}" AS VARCHAR), '\\(([^)]+)\\)', 1)))
                FROM {table_name}
                WHERE "{col}" IS NOT NULL
                  AND REGEXP_EXTRACT(CAST("{col}" AS VARCHAR), '\\(([^)]+)\\)', 1) != ''
            """).fetchone()[0]
            if code_in_parens_len:
                clean_len = code_in_parens_len
            geo_type = None
            if 4.5 <= clean_len <= 5.5:
                geo_type = "Code INSEE ou commune (zéros perdus)" if is_numeric else "Code INSEE ou commune"
            elif 3.5 <= clean_len <= 4.5:
                geo_type = "Code INSEE ou commune (zéros perdus)"
            elif 1.5 <= clean_len <= 3.5:
                if 'reg' in w_match:
                    geo_type = "Code region"
                elif 'dep' in w_match or 'departement' in w_match or 'département' in w_match:
                    geo_type = "Code departement (extrait)"
                elif 'com' in w_match or 'commune' in w_match:
                    geo_type = "Code INSEE ou commune (extrait)"
                else:
                    geo_type = "Code region ou departement"
            if geo_type:
                candidates.append({"col": col, "label": geo_type})
        except Exception:
            continue
    return candidates

def format_geo_keys_table(geo_keys):
    if not geo_keys:
        return "None"
    return {
        "_table": True,
        "data": [
            {"Reference area": k["label"], "Identified key": k["col"]}
            for k in geo_keys
        ]
    }
    
def get_geo_columns_duckdb(conn, table_name):
    """Identify geometry columns using DuckDB queries."""
    lat_pattern = r'\b(lat|latitude)\b'
    lon_pattern = r'\b(lon|long|lng|longitude)\b'
    x_pattern = r'(^x|[^a-zA-Z0-9]x$)'
    y_pattern = r'(^y|[^a-zA-Z0-9]y$)'
    geom_pattern = r'\b(geometry|geom|shape|point|polygon)\b'
    addr_pattern = r'adresse'
    insee_pattern = r'(dep|reg|insee|com)'

    result = {
        'columns': [],
        'type': None,
        'method': None,
        'geo_keys': [],
        'geotrans': 'Aucune geometry',
    }

    schema = conn.execute(f"DESCRIBE {table_name}").fetchdf()
    columns = schema['column_name'].tolist()

    # Check for geometry columns (GeoJSON/WKT)
    for col in columns:
        col_lc = col.lower()
        if not re.search(geom_pattern, col_lc):
            continue

        try:
            sample = conn.execute(f"""
                SELECT "{col}" FROM {table_name}
                WHERE "{col}" IS NOT NULL
                LIMIT 1
            """).fetchone()

            if sample is None:
                continue

            sample_val = sample[0]

            # Check geo_point format
            if 'point' in col_lc and 'geo' in col_lc and isinstance(sample_val, str) and ',' in sample_val:
                return {**result, 'type': 'Point', 'method': 'geopoint', 'columns': [col], 'geotrans': "Présence Géométrie"}

            # Check GeoJSON
            try:
                val = json.loads(sample_val) if isinstance(sample_val, str) else sample_val
                if isinstance(val, dict) and 'type' in val and 'coordinates' in val:
                    return {**result, 'type': val.get('type', 'Unknown'), 'method': 'geojson', 'columns': [col], 'geotrans': "Présence Géométrie"}
            except Exception:
                pass

            # Check WKT
            if isinstance(sample_val, str):
                sample_upper = sample_val.upper()
                for geom_type in ['POINT', 'LINESTRING', 'POLYGON']:
                    if geom_type in sample_upper:
                        return {**result, 'type': geom_type.title(), 'method': 'from_wkt', 'columns': [col], 'geotrans': "Présence Géométrie"}
        except Exception:
            continue

    # Check for Lat/Lon pairs
    lat_col = lon_col = None
    for col in columns:
        col_lc = col.lower()
        if re.search(lat_pattern, col_lc) and lat_col is None:
            lat_col = col
        if re.search(lon_pattern, col_lc) and lon_col is None:
            lon_col = col

    if lat_col and lon_col:
        return {**result, 'type': 'Point', 'method': 'points_from_xy', 'columns': [lon_col, lat_col], 'geotrans': "Présence géométrie séparée (x,y)"}

    # Check for X/Y pairs
    x_cols = [col for col in columns if re.search(x_pattern, col.lower())]
    y_cols = [col for col in columns if re.search(y_pattern, col.lower())]

    if x_cols and y_cols:
        # Match X/Y pairs by suffix and check for start/end coordinate patterns (LineString)
        for x_col in x_cols:
            for y_col in y_cols:
                x_suffix = x_col.lower().replace('x', '')
                y_suffix = y_col.lower().replace('y', '')

                if x_suffix == y_suffix or (not x_suffix and not y_suffix):
                    # Check if it's start/end coordinates for LineString
                    if any(k in x_col.lower() for k in ['d', 'debut', 'start']):
                        for x_end in x_cols:
                            if x_end != x_col and any(k in x_end.lower() for k in ['f', 'fin', 'end']):
                                for y_end in y_cols:
                                    if y_end != y_col and any(k in y_end.lower() for k in ['f', 'fin', 'end']):
                                        return {**result, 'type': 'LineString', 'method': 'linestring_coords',
                                                'columns': [x_col, y_col, x_end, y_end],
                                                'geotrans': "Présence géométrie multiples (x1,y1), (x2,y2)"}

                    # Otherwise, simple Point pair
                    return {**result, 'type': 'Point', 'method': 'points_from_xy', 'columns': [x_col, y_col], 'geotrans': "Présence géométrie séparée (x,y)"}

    # Check for address columns
    for col in columns:
        if re.search(addr_pattern, col.lower()):
            return {**result, 'type': 'Address', 'method': 'geocoding_required', 'columns': [col], 'geotrans': "Géocodage de l'adresse"}

    # Collect INSEE/geographic keys
    for col in columns:
        if re.search(insee_pattern, col.lower()):
            result['geo_keys'].append(col)

    if result['geo_keys']:
        result['type'] = 'Administrative'
        result['method'] = 'join_required'
        result['columns'] = result['geo_keys']
        result['geotrans'] = "Jointure spatiale à l'aide de clés géographiques"

    return result


# ============================================================================
# DUCKDB CSV INSPECTION
# ============================================================================
def guess_crs_from_coords_duckdb(conn, table_name, x_col, y_col):
    """
    Guess CRS from coordinate columns using score-based detection.
    Fetches a sample and delegates to guess_crs_from_xy.
    """
    try:
        df = conn.execute(f"""
            SELECT "{x_col}" AS x, "{y_col}" AS y
            FROM {table_name}
            WHERE "{x_col}" IS NOT NULL AND "{y_col}" IS NOT NULL
        """).fetchdf()
        if df.empty:
            return None
        epsg, conf, alts = guess_crs_from_xy(df["x"].values, df["y"].values)
        if epsg is None:
            epsg = alts[0] if alts else None
            print(f"[CRS] Low-confidence detection — using best candidate EPSG:{epsg}. Alternatives: {alts}")
        else:
            print(f"[CRS] Detected EPSG:{epsg} (confidence={conf:.2f})")
        return epsg
    except Exception as e:
        print(f"[CRS] guess_crs_from_coords_duckdb error: {e}")
        return None

# ============================================================================
# SHARED PROJ STRINGS
# ============================================================================
PROJ_STRINGS = {
    4326:  '+proj=longlat +datum=WGS84',
    2154:  '+proj=lcc +lat_0=46.5 +lon_0=3 +lat_1=49 +lat_2=44 +x_0=700000 +y_0=6600000 +ellps=GRS80 +units=m',
    3857:  '+proj=merc +a=6378137 +b=6378137 +lat_ts=0 +lon_0=0 +x_0=0 +y_0=0 +k=1 +units=m',
    27700: '+proj=tmerc +lat_0=49 +lon_0=-2 +k=0.9996012717 +x_0=400000 +y_0=-100000 +ellps=airy +units=m',
    25832: '+proj=utm +zone=32 +ellps=GRS80 +units=m',
    3035:  '+proj=laea +lat_0=52 +lon_0=10 +x_0=4321000 +y_0=3210000 +ellps=GRS80 +units=m',
    5070:  '+proj=aea +lat_0=23 +lon_0=-96 +lat_1=29.5 +lat_2=45.5 +x_0=0 +y_0=0 +datum=NAD83 +units=m',
    2062:  '+proj=lcc +lat_1=40 +lat_2=40 +lat_0=40 +lon_0=-3 +x_0=2500000 +y_0=0 +ellps=intl +units=m',
}

# ============================================================================
# CRS DETECTION — scoring-based approach
# ============================================================================
    
# Loose area-of-use boxes: (min_lon, min_lat, max_lon, max_lat)
_CANDIDATE_CRS = {
    4326:  (-180, -90,  180,  90),
    3857:  (-180, -85,  180,  85),
    2154:  (  -6,  41,   10,  52),   # Lambert 93
    27700: (  -9,  49,    3,  62),   # OSGB / British National Grid
    3035:  ( -35,  25,   45,  85),   # ETRS89-LAEA Europe
    25830: ( -12,  34,    0,  48),
    25831: (  -6,  34,    6,  48),
    25832: (   0,  34,   12,  84),
    25833: (   6,  34,   18,  84),
    32630: ( -12,   0,    0,  84),
    32631: (  -6,   0,    6,  84),
    32632: (   0,   0,   12,  84),
    32633: (   6,   0,   18,  84),
    5070:  (-180,  15, -50,  75),    # NAD83 Albers (US)
    2062:  (  -9,  36,    5,  44),   # Madrid 1870 (Spain)
}


def _compute_xy_stats(xs, ys):
    return {
        "median_x":    float(np.median(xs)),
        "median_y":    float(np.median(ys)),
        "min_x":       float(xs.min()),  "max_x": float(xs.max()),
        "min_y":       float(ys.min()),  "max_y": float(ys.max()),
        "range_x":     float(xs.max() - xs.min()),
        "range_y":     float(ys.max() - ys.min()),
    }


def _looks_like_degrees(st):
    return (
        -200 <= st["min_x"] and st["max_x"] <= 200 and
        -100 <= st["min_y"] and st["max_y"] <= 100 and
        st["range_x"] <= 60 and st["range_y"] <= 60
    )


def _looks_like_projected(st):
    return (
        1_000 <= abs(st["median_x"]) <= 20_000_000 and
        1_000 <= abs(st["median_y"]) <= 20_000_000
    )


def _score_candidate(epsg, xs, ys, max_points=500):
    from pyproj import Transformer, CRS as ProjCRS
    if len(xs) > max_points:
        idx = np.random.choice(len(xs), size=max_points, replace=False)
        xs, ys = xs[idx], ys[idx]
    try:
        t = Transformer.from_crs(ProjCRS.from_epsg(epsg), ProjCRS.from_epsg(4326), always_xy=True)
        lon, lat = t.transform(xs, ys)
        lon, lat = np.asarray(lon), np.asarray(lat)
    except Exception:
        return -1e9
    mask = np.isfinite(lon) & np.isfinite(lat) & (lon >= -180) & (lon <= 180) & (lat >= -90) & (lat <= 90)
    lon, lat = lon[mask], lat[mask]
    if lon.size == 0:
        return -1e9
    bbox_area = (lon.max() - lon.min()) * (lat.max() - lat.min())
    area_score = 1.0 / (1.0 + bbox_area)
    aoa = _CANDIDATE_CRS.get(epsg)
    if aoa:
        in_region = (lon.min() >= aoa[0] - 5 and lon.max() <= aoa[2] + 5 and
                     lat.min() >= aoa[1] - 5 and lat.max() <= aoa[3] + 5)
        region_score = 1.0 if in_region else 0.2
    else:
        region_score = 0.5
    return 0.7 * region_score + 0.3 * area_score


def guess_crs_from_xy(xs, ys, min_conf_gap=0.15):
    """
    Score-based CRS detection from raw coordinate arrays.
    Returns (epsg: int | None, confidence: float, alternatives: list[int])
    """
    xs, ys = np.asarray(xs, dtype=float), np.asarray(ys, dtype=float)
    xs, ys = xs[np.isfinite(xs)], ys[np.isfinite(ys)]
    if xs.size == 0:
        return None, 0.0, []

    st = _compute_xy_stats(xs, ys)

    if _looks_like_degrees(st):
        return 4326, 1.0, [4326]

    if not _looks_like_projected(st):
        return None, 0.0, []

    candidates = list(_CANDIDATE_CRS.keys())
    scores = {epsg: _score_candidate(epsg, xs, ys) for epsg in candidates}
    ordered = sorted(scores.items(), key=lambda kv: kv[1], reverse=True)
    best_epsg, best_score = ordered[0]
    second_score = ordered[1][1] if len(ordered) > 1 else -1e9

    if best_score - second_score < min_conf_gap or best_score <= 0:
        return None, float(best_score), [e for e, _ in ordered[:3]]

    return int(best_epsg), float(best_score), [e for e, _ in ordered[:3]]




def process_geometry_duckdb_points(conn, table_name, x_col, y_col, gdf_metro, wgs84_bounds=None, metric_crs=None):
    """
    Process point geometry (from x,y columns) entirely in DuckDB using spatial functions.
    Filters using bounds (transformed to detected CRS if needed).
    Reprojects to metric_crs if provided, otherwise keeps source CRS.
    """
    print(f"[DuckDB Spatial] Processing point geometry with DuckDB spatial extension...")
    start = time.time()

    DEFAULT_BOUNDS = [-5.5, 41.0, 10.0, 51.5]  # France fallback (WGS84)
    bounds = wgs84_bounds or DEFAULT_BOUNDS
    minx, miny, maxx, maxy = bounds

    # Detect CRS from coordinate values
    detected_crs = guess_crs_from_coords_duckdb(conn, table_name, x_col, y_col)
    if detected_crs is None:
        print(f"[DuckDB Spatial] No CRS detected, keeping as-is")
    else:
        print(f"[DuckDB Spatial] Detected CRS: EPSG:{detected_crs}")

    # Transform bounds to detected CRS if not WGS84
    if detected_crs and detected_crs != 4326:
        from pyproj import Transformer
        transformer = Transformer.from_crs("EPSG:4326", f"EPSG:{detected_crs}", always_xy=True)
        minx_t, miny_t = transformer.transform(minx, miny)
        maxx_t, maxy_t = transformer.transform(maxx, maxy)
        print(f"[DuckDB Spatial] Filtering to bounds (EPSG:{detected_crs}): x=[{minx_t:.0f},{maxx_t:.0f}] y=[{miny_t:.0f},{maxy_t:.0f}]")
    else:
        minx_t, miny_t, maxx_t, maxy_t = minx, miny, maxx, maxy
        print(f"[DuckDB Spatial] Filtering to bounds (WGS84): lon=[{minx},{maxx}] lat=[{miny},{maxy}]")

    conn.execute(f"""
        CREATE TABLE geo_filtered_raw AS
        SELECT * FROM {table_name}
        WHERE "{x_col}" IS NOT NULL AND "{y_col}" IS NOT NULL
          AND "{x_col}" BETWEEN {minx_t} AND {maxx_t}
          AND "{y_col}" BETWEEN {miny_t} AND {maxy_t}
    """)

    total_with_coords = conn.execute(f"""
        SELECT COUNT(*) FROM {table_name}
        WHERE "{x_col}" IS NOT NULL AND "{y_col}" IS NOT NULL
    """).fetchone()[0]

    filtered_count = conn.execute("SELECT COUNT(*) FROM geo_filtered_raw").fetchone()[0]
    if filtered_count < total_with_coords:
        excluded = total_with_coords - filtered_count
        print(f"[DuckDB Spatial] Filtered to bounds: {filtered_count:,} points ({excluded:,} excluded)")

    # No reprojection — keep geometry in source CRS (WGS84 for most international datasets)
    # Reprojection would break coverage calculation (gdf_metro is in EPSG:2154) and map display
    ## if metric_crs and detected_crs and detected_crs != metric_crs:
    ##     source_proj = PROJ_STRINGS.get(detected_crs, f'EPSG:{detected_crs}')
    ##     target_proj = PROJ_STRINGS.get(metric_crs, f'EPSG:{metric_crs}')
    ##     print(f"[DuckDB Spatial] Transforming from EPSG:{detected_crs} to EPSG:{metric_crs}...")
    ##     conn.execute(f"""
    ##         CREATE TABLE geo_processed AS
    ##         SELECT *,
    ##             ST_Transform(ST_Point("{x_col}", "{y_col}"), '{source_proj}', '{target_proj}') as geom
    ##         FROM geo_filtered_raw
    ##     """)
    ## else:
    print(f"[DuckDB Spatial] No reprojection — keeping EPSG:{detected_crs}")
    conn.execute(f"""
        CREATE TABLE geo_processed AS
        SELECT *, ST_Point("{x_col}", "{y_col}") as geom
        FROM geo_filtered_raw
    """)

    conn.execute("DROP TABLE geo_filtered_raw")

    return _compute_spatial_metrics_duckdb(conn, "geo_processed", "Point", gdf_metro, start, source_crs=detected_crs)


def process_geometry_duckdb_linestrings(conn, table_name, x_start, y_start, x_end, y_end, gdf_metro, wgs84_bounds=None, metric_crs=None):
    """
    Process LineString geometry from start/end coordinate columns using DuckDB spatial.
    Uses ST_MakeLine to build LineStrings from (xD,yD)→(xF,yF) pairs.
    Handles French decimal separators (comma) via REPLACE normalization.
    Filters using bounds (transformed to detected CRS if needed).
    """
    print(f"[DuckDB Spatial] Processing LineString geometry with DuckDB spatial extension...")
    start = time.time()

    DEFAULT_BOUNDS = [-5.5, 41.0, 10.0, 51.5]
    bounds = wgs84_bounds or DEFAULT_BOUNDS
    minx, miny, maxx, maxy = bounds

    # Safe cast: handles both numeric columns and VARCHAR with comma decimal separator
    def safe_cast(col):
        return f"TRY_CAST(REPLACE(CAST(\"{col}\" AS VARCHAR), ',', '.') AS DOUBLE)"

    # Materialize normalized coordinates
    conn.execute(f"""
        CREATE TABLE geo_coords AS
        SELECT
            {safe_cast(x_start)} as x_s,
            {safe_cast(y_start)} as y_s,
            {safe_cast(x_end)} as x_e,
            {safe_cast(y_end)} as y_e
        FROM {table_name}
        WHERE {safe_cast(x_start)} IS NOT NULL AND {safe_cast(y_start)} IS NOT NULL
          AND {safe_cast(x_end)} IS NOT NULL AND {safe_cast(y_end)} IS NOT NULL
    """)

    # Detect CRS from start coordinates BEFORE filtering
    detected_crs = guess_crs_from_coords_duckdb(conn, "geo_coords", "x_s", "y_s")
    if detected_crs is None:
        print(f"[DuckDB Spatial] No CRS detected, keeping as-is")
    else:
        print(f"[DuckDB Spatial] Detected CRS: EPSG:{detected_crs}")

    # Transform bounds to detected CRS if not WGS84
    if detected_crs and detected_crs != 4326:
        from pyproj import Transformer
        transformer = Transformer.from_crs("EPSG:4326", f"EPSG:{detected_crs}", always_xy=True)
        minx_t, miny_t = transformer.transform(minx, miny)
        maxx_t, maxy_t = transformer.transform(maxx, maxy)
        print(f"[DuckDB Spatial] Filtering to bounds (EPSG:{detected_crs}): x=[{minx_t:.0f},{maxx_t:.0f}] y=[{miny_t:.0f},{maxy_t:.0f}]")
    else:
        minx_t, miny_t, maxx_t, maxy_t = minx, miny, maxx, maxy
        print(f"[DuckDB Spatial] Filtering to bounds (WGS84): lon=[{minx},{maxx}] lat=[{miny},{maxy}]")

    # Filter on bounds (now in correct CRS)
    conn.execute(f"""
        CREATE TABLE geo_coords_filtered AS
        SELECT * FROM geo_coords
        WHERE x_s BETWEEN {minx_t} AND {maxx_t}
          AND y_s BETWEEN {miny_t} AND {maxy_t}
    """)
    conn.execute("DROP TABLE geo_coords")
    conn.execute("ALTER TABLE geo_coords_filtered RENAME TO geo_coords")

    # No reprojection — keep geometry in source CRS
    ## if metric_crs and detected_crs and detected_crs != metric_crs:
    ##     source_proj = PROJ_STRINGS.get(detected_crs, f'EPSG:{detected_crs}')
    ##     target_proj = PROJ_STRINGS.get(metric_crs, f'EPSG:{metric_crs}')
    ##     print(f"[DuckDB Spatial] Transforming from EPSG:{detected_crs} to EPSG:{metric_crs}...")
    ##     conn.execute(f"""
    ##         CREATE TABLE geo_lines AS
    ##         SELECT *,
    ##             ST_Transform(
    ##                 ST_MakeLine(ST_Point(x_s, y_s), ST_Point(x_e, y_e)),
    ##                 '{source_proj}', '{target_proj}'
    ##             ) as geom
    ##         FROM geo_coords
    ##     """)
    ## else:
    print(f"[DuckDB Spatial] No reprojection — keeping EPSG:{detected_crs}")
    conn.execute(f"""
        CREATE TABLE geo_lines AS
        SELECT *,
            ST_MakeLine(ST_Point(x_s, y_s), ST_Point(x_e, y_e)) as geom
        FROM geo_coords
    """)

    conn.execute("DROP TABLE IF EXISTS geo_coords")

    # geo_lines becomes geo_processed directly — no second spatial filter needed
    # (filtering already done on raw coords above)
    conn.execute("ALTER TABLE geo_lines RENAME TO geo_processed")

    return _compute_spatial_metrics_duckdb(conn, "geo_processed", "LineString", gdf_metro, start, source_crs=detected_crs)


def guess_crs_from_bounds_duckdb(conn, table_name, geom_col):
    """
    Guess CRS from a geometry column using score-based detection.
    Extracts centroid coordinates and delegates to guess_crs_from_xy.
    Only used as fallback when CRS metadata is absent from the file.
    """
    try:
        df = conn.execute(f"""
            SELECT
                ST_X(ST_Centroid("{geom_col}")) AS x,
                ST_Y(ST_Centroid("{geom_col}")) AS y
            FROM {table_name}
            WHERE "{geom_col}" IS NOT NULL
        """).fetchdf()
        if df.empty:
            return None
        epsg, conf, alts = guess_crs_from_xy(df["x"].values, df["y"].values)
        if epsg is None:
            print(f"[CRS] Low-confidence detection. Top candidates: {alts}")
        else:
            print(f"[CRS] Detected EPSG:{epsg} (confidence={conf:.2f})")
        return epsg
    except Exception as e:
        print(f"[CRS] guess_crs_from_bounds_duckdb error: {e}")
        return None


def process_geometry_duckdb_native(conn, table_name, geom_col, gdf_metro, wgs84_bounds=None, metric_crs=None):
    """
    Process native geometry column (from geospatial files) using DuckDB spatial.
    Handles Points, LineStrings, Polygons, etc.
    Filters using WGS84 bounds when source CRS is 4326.
    Reprojects to metric_crs if provided, otherwise keeps source CRS.
    """
    print(f"[DuckDB Spatial] Processing native geometry with DuckDB spatial extension...")
    start = time.time()

    DEFAULT_BOUNDS = [-5.5, 41.0, 10.0, 51.5]  # France fallback
    bounds = wgs84_bounds or DEFAULT_BOUNDS
    minx, miny, maxx, maxy = bounds

    # Get geometry type
    geom_type_result = conn.execute(f"""
        SELECT ST_GeometryType("{geom_col}") FROM {table_name}
        WHERE "{geom_col}" IS NOT NULL LIMIT 1
    """).fetchone()
    geom_type = geom_type_result[0] if geom_type_result else "GEOMETRY"

    try:
        crs_result = conn.execute(f"""
            SELECT ST_SRID("{geom_col}") FROM {table_name}
            WHERE "{geom_col}" IS NOT NULL LIMIT 1
        """).fetchone()
        detected_crs = crs_result[0] if crs_result and crs_result[0] != 0 else None
    except Exception:
        detected_crs = None
    
    if detected_crs:
        print(f"[DuckDB Spatial] CRS read from file metadata: EPSG:{detected_crs}")
    else:
        # Fallback for files with missing CRS metadata
        detected_crs = guess_crs_from_bounds_duckdb(conn, table_name, geom_col)
        print(f"[DuckDB Spatial] CRS guessed from coordinates: EPSG:{detected_crs}")

    # No reprojection — keep geometry in source CRS
    ## if metric_crs and detected_crs and detected_crs != metric_crs:
    ##     source_proj = PROJ_STRINGS.get(detected_crs, f'EPSG:{detected_crs}')
    ##     target_proj = PROJ_STRINGS.get(metric_crs, f'EPSG:{metric_crs}')
    ##     print(f"[DuckDB Spatial] Transforming from EPSG:{detected_crs} to EPSG:{metric_crs}...")
    ##     conn.execute(f"""
    ##         CREATE TABLE geo_transformed AS
    ##         SELECT *,
    ##             ST_Transform("{geom_col}", '{source_proj}', '{target_proj}') as geom_target
    ##         FROM {table_name}
    ##         WHERE "{geom_col}" IS NOT NULL
    ##     """)
    ##     work_geom_col = "geom_target"
    ##     work_table = "geo_transformed"
    ## else:
    print(f"[DuckDB Spatial] No reprojection — keeping EPSG:{detected_crs}")
    conn.execute(f"""
        CREATE TABLE geo_transformed AS
        SELECT *, "{geom_col}" as geom_target
        FROM {table_name}
        WHERE "{geom_col}" IS NOT NULL
    """)
    work_geom_col = "geom_target"
    work_table = "geo_transformed"

    # Create processed table with centroid for filtering
    conn.execute(f"""
        CREATE TABLE geo_processed AS
        SELECT * EXCLUDE ("{geom_col}"),
            CASE
                WHEN ST_GeometryType({work_geom_col}) = 'POINT' THEN {work_geom_col}
                ELSE ST_Centroid({work_geom_col})
            END as geom_center,
            {work_geom_col} as geom
        FROM {work_table}
    """)

    # Filter using WGS84 bounds only when source CRS is 4326
    # (centroid is still in WGS84 before reprojection in this case)
    if detected_crs == 4326:
        conn.execute(f"""
            CREATE TABLE geo_filtered AS
            SELECT * FROM geo_processed
            WHERE ST_X(geom_center) BETWEEN {minx} AND {maxx}
              AND ST_Y(geom_center) BETWEEN {miny} AND {maxy}
        """)
    else:
        # For metric CRS sources, bounds don't apply — keep all
        conn.execute("CREATE TABLE geo_filtered AS SELECT * FROM geo_processed")

    filtered_count = conn.execute("SELECT COUNT(*) FROM geo_filtered").fetchone()[0]
    total_count = conn.execute("SELECT COUNT(*) FROM geo_processed").fetchone()[0]

    if filtered_count < total_count:
        excluded = total_count - filtered_count
        print(f"[DuckDB Spatial] Filtered to bounds: {filtered_count:,} geometries ({excluded:,} excluded)")

    conn.execute("DROP TABLE IF EXISTS geo_processed")
    conn.execute("DROP TABLE IF EXISTS geo_transformed")
    conn.execute("ALTER TABLE geo_filtered RENAME TO geo_processed")

    return _compute_spatial_metrics_duckdb(conn, "geo_processed", geom_type, gdf_metro, start, source_crs=detected_crs)

    
def _compute_spatial_metrics_duckdb(conn, table_name, geom_type, gdf_metro, start_time, source_crs=None):
    """
    Compute spatial metrics using DuckDB spatial functions.
    Shared logic for both point and native geometry processing.
    """
    total = conn.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]

    if total == 0:
        return get_default_geo_summary()

    non_empty = total  # All rows have geometry by construction

    # Validate geometries
    valid_count = conn.execute(f"SELECT COUNT(*) FROM {table_name} WHERE ST_IsValid(geom)").fetchone()[0]

    hull_row = conn.execute(f"""
    SELECT
        ST_AsText(ST_Envelope_Agg(geom)) AS hull_wkt
    FROM {table_name}
    """).fetchone()

    hull_wkt = hull_row[0] if hull_row and hull_row[0] is not None else None

    # Compute area in km² using a metric CRS — DuckDB area is in source CRS units
    # which are degrees² for WGS84, making it meaningless without reprojection
    hull_area_km2 = 0
    if hull_wkt:
        try:
            from shapely import wkt as shapely_wkt
            hull_geom_raw = shapely_wkt.loads(hull_wkt)
            # Pick an appropriate metric CRS for the area calculation
            # Use the source CRS if already metric, otherwise reproject to a suitable one
            METRIC_CRS_FOR_AREA = {
                4326:  3857,   # WGS84 → Web Mercator (good enough for area estimates)
                27700: 27700,  # British National Grid — already metric
                25832: 25832,  # UTM 32N — already metric
                3035:  3035,   # ETRS89-LAEA — already metric
                5070:  5070,   # NAD83 Albers — already metric
                2062:  2062,   # Madrid — already metric
                2154:  2154,   # Lambert 93 — already metric
            }
            area_crs = METRIC_CRS_FOR_AREA.get(source_crs, 3857) if source_crs else 3857
            if source_crs and source_crs != area_crs:
                hull_gs = gpd.GeoSeries([hull_geom_raw], crs=f"EPSG:{source_crs}").to_crs(epsg=area_crs)
                hull_area_km2 = hull_gs.iloc[0].area / 1e6
            else:
                hull_area_km2 = hull_geom_raw.area / 1e6
        except Exception as e:
            print(f"[DuckDB Spatial] Hull area calculation error: {e}")
            hull_area_km2 = 0

    # Compute bounding box
    bbox = conn.execute(f"""
    SELECT
        MIN(ST_X(ST_Centroid(geom))) as minx, MIN(ST_Y(ST_Centroid(geom))) as miny,
        MAX(ST_X(ST_Centroid(geom))) as maxx, MAX(ST_Y(ST_Centroid(geom))) as maxy
    FROM {table_name}""").fetchone()

    minx, miny, maxx, maxy = bbox

    if 'POINT' in geom_type.upper():
        area_km2 = hull_area_km2
        complexity = "None : POINT"

    elif 'LINESTRING' in geom_type.upper() or 'LINE' in geom_type.upper():
        area_km2 = hull_area_km2
        complexity_result = conn.execute(f"""
            SELECT AVG(ST_NPoints(geom)) FROM {table_name}
        """).fetchone()
        complexity = round(complexity_result[0], 2) if complexity_result[0] else 0

    else:
        try:
            area_result = conn.execute(f"""
                SELECT ST_Area(ST_Union_Agg(geom)) / 1e6 as area_km2
                FROM {table_name}
            """).fetchone()
            area_km2 = area_result[0] if area_result[0] else hull_area_km2
        except Exception:
            area_km2 = hull_area_km2

        complexity_result = conn.execute(f"""
            SELECT AVG(ST_NPoints(geom)) FROM {table_name}
        """).fetchone()
        complexity = round(complexity_result[0], 2) if complexity_result[0] else 0

    # Density
    density = total / area_km2 if area_km2 > 0 else 0

    # Duplicate detection using WKT
    dup_result = conn.execute(f"""
        WITH wkt_geoms AS (
            SELECT ST_AsText(geom) as wkt FROM {table_name}
        )
        SELECT
            COUNT(*) as total,
            COUNT(DISTINCT wkt) as unique_count
        FROM wkt_geoms
    """).fetchone()
    dup_pct = (dup_result[0] - dup_result[1]) / dup_result[0] * 100 if dup_result[0] > 0 else 0

    # bbox_area must use the same metric hull for consistency
    # Re-use hull_area_km2 as the bbox approximation (envelope already is the bbox)
    bbox_area = hull_area_km2  # hull is the envelope
    fill_rate = (area_km2 / bbox_area * 100) if bbox_area > 0 else 0

    # Compute coverage against reference using precomputed hull_wkt
    # Only meaningful when data and reference share the same CRS (both EPSG:2154 for France)
    coverage_pct = 0
    try:
        if hull_wkt and gdf_metro is not None:
            from shapely import wkt as shapely_wkt
            from shapely.validation import make_valid
            ref_crs = gdf_metro.crs.to_epsg() if gdf_metro.crs else None
            # Skip coverage if data CRS differs from reference — intersection would be meaningless
            if source_crs is None or ref_crs is None or source_crs == ref_crs:
                hull_geom = make_valid(shapely_wkt.loads(hull_wkt))
                ref_dissolved = make_valid(gdf_metro.union_all())
                try:
                    intersection = hull_geom.intersection(ref_dissolved)
                    coverage_pct = (intersection.area / ref_dissolved.area) * 100 if ref_dissolved.area > 0 else 0
                except Exception:
                    coverage_pct = 0
            else:
                print(f"[DuckDB Spatial] Coverage skipped — data CRS (EPSG:{source_crs}) differs from reference CRS (EPSG:{ref_crs})")
    except Exception as e:
        print(f"[DuckDB Spatial] Coverage calculation error: {e}")

    processing_time = time.time() - start_time
    print(f"[DuckDB Spatial] Geometry processing done in {processing_time:.2f}s")

    # Clean geometry type for display
    display_geom_type = geom_type.replace("ST_", "").title()

    return {
        "Score de complétude géographique": {
            "_table": True,
            "data": [{
                "Présentes (%)": round(non_empty / total * 100, 1),
                "Valides (%)":   round(valid_count / total * 100, 1)}]},
        "CRS": f"EPSG:{source_crs}" if source_crs else "Unknown",
        "Types de géométrie": display_geom_type,
        "Emprise estimée (km2)": round(area_km2, 2),
        "Densité (obj/km2)": round(density, 2),
        "Taux de remplissage géométrique (%)": round(fill_rate, 2),
        "Complexité moyenne des géométries": complexity,
        "Part des geometries dupliquees (%)": round(dup_pct, 2),
        "Couverture territoriale (%)": round(coverage_pct, 2),
    }


def process_geometry_duckdb(conn, table_name, x_col, y_col, gdf_metro, wgs84_bounds=None, metric_crs=None):
    """Backward compatible wrapper for point geometry processing."""
    return process_geometry_duckdb_points(conn, table_name, x_col, y_col, gdf_metro, wgs84_bounds=wgs84_bounds, metric_crs=metric_crs)


def inspect_csv_duckdb(filepath, gdf_metro, geo_key_patterns=None, wgs84_bounds=None, metric_crs=None):
    """
    Inspect CSV file using DuckDB for faster processing.
    """
    global last_gdf

    print(f"[DuckDB] Inspecting CSV: {filepath}")
    start_time = time.time()

    meta = get_file_metadata(filepath)

    # Create DuckDB connection
    conn = get_duckdb_connection()

    # Read CSV directly with DuckDB (auto-detects format)
    try:
        conn.execute(f"""
            CREATE TABLE csv_data AS
            SELECT * FROM read_csv('{filepath}',
                header=true,
                auto_detect=true,
                ignore_errors=true,
                sample_size=100000
            )
        """)
    except Exception as e:
        print(f"[DuckDB] Error reading CSV: {e}")
        conn.execute(f"""
            CREATE TABLE csv_data AS
            SELECT * FROM read_csv('{filepath}',
                header=true,
                all_varchar=true,
                ignore_errors=true
            )
        """)

    read_time = time.time() - start_time
    print(f"[DuckDB] CSV read in {read_time:.2f}s")
    
    # ── Detect and fix misparsed date columns ────────────────────────────
    # Strategy: read raw file with pandas to find ALL date-like columns,
    # then force them all to VARCHAR in DuckDB from the start.
    bad_date_cols = []
    try:
        with open(filepath, 'rb') as _f:
            _first = _f.read(2048).decode('utf-8', errors='replace')
        _sep = ';' if _first.count(';') > _first.count(',') else ','
        df_probe = pd.read_csv(filepath, nrows=3, dtype=str,
                               sep=_sep, encoding='utf-8-sig', encoding_errors='replace')
        for col in df_probe.columns:
            vals = df_probe[col].dropna()
            if vals.empty:
                continue
            s = str(vals.iloc[0]).strip()
            if re.match(r'^\d{2}/\d{2}/\d{4}', s) or \
               re.match(r'^\d{2}-\d{2}-\d{4}', s):
                bad_date_cols.append(col)
    except Exception as e:
        print(f"[DateTime] Probe read failed: {e}")

    if bad_date_cols:
        print(f"[DateTime] Ambiguous date columns detected: {bad_date_cols} — forcing VARCHAR")
        col_types = ", ".join([f'"{c}": "VARCHAR"' for c in bad_date_cols])
        conn.execute("DROP TABLE IF EXISTS csv_data")
        try:
            conn.execute(f"""
                CREATE TABLE csv_data AS
                SELECT * FROM read_csv('{filepath}',
                    header=true,
                    auto_detect=true,
                    ignore_errors=true,
                    sample_size=100000,
                    types={{{col_types}}}
                )
            """)
            print(f"[DateTime] Re-read done, forced VARCHAR for: {bad_date_cols}")
        except Exception as e:
            print(f"[DateTime] Re-read failed: {e}")
            conn.execute(f"""
                CREATE TABLE csv_data AS
                SELECT * FROM read_csv('{filepath}',
                    header=true,
                    all_varchar=true,
                    ignore_errors=true
                )
            """)
    # ─────────────────────────────────────────────────────────────────────
    # ────────────────────────────────────────────────────────

    # Get row and column counts
    row_count = conn.execute("SELECT COUNT(*) FROM csv_data").fetchone()[0]
    col_count = len(conn.execute("DESCRIBE csv_data").fetchdf())

    # Detect geo info
    geo_keys = detect_geo_join_keys_duckdb(conn, "csv_data", geo_key_patterns=geo_key_patterns)
    res_geom = get_geo_columns_duckdb(conn, "csv_data")

    geo_key_cols = [k["col"] for k in geo_keys]
    geo_key_completeness = completeness_score_duckdb_cols(conn, "csv_data", geo_key_cols)

    # Detect temporal/date columns (preliminary - full analysis done by analyze_datetime_columns)
    date_cols_detected = detect_date_columns_duckdb(conn, "csv_data")
    if date_cols_detected:
        print(f"[DuckDB] Date columns detected: {date_cols_detected}")

    # Build base summary
    base_summary = {
        **meta,
        "Type de fichier": "CSV (DuckDB)",
        "Nb lignes": row_count,
        "Nb colonnes": col_count,
        "Colonnes": {"_table": True, "data": build_columns_detail_duckdb(conn, "csv_data")},
        "Score de complétude global": completeness_score_duckdb(conn, "csv_data"),
        "Analyse temporelle": None,  # Will be populated by analyze_datetime_columns
        "Clés géographiques": format_geo_keys_table(geo_keys),
        "Géotransformation": res_geom['geotrans'],
        "Score de complétude des clés géographique": geo_key_completeness,
    }

    geo_summary = get_default_geo_summary()

    # Process geometry if detected
    if res_geom['columns'] and res_geom['method'] == 'points_from_xy':
        print(f"[DuckDB] Geometry detected: {res_geom['columns']} ({res_geom['method']})")

        x_col, y_col = res_geom['columns']

        geo_summary = process_geometry_duckdb(conn, "csv_data", x_col, y_col, gdf_metro,
                                               wgs84_bounds=wgs84_bounds, metric_crs=metric_crs)
        base_summary["Type de fichier"] = "CSV with Geometry (DuckDB Spatial)"

        # Create GeoDataFrame for map display — sample directly from csv_data within bounds
        try:
            bounds = wgs84_bounds or [-5.5, 41.0, 10.0, 51.5]
            bminx, bminy, bmaxx, bmaxy = bounds
            detected_crs_val = guess_crs_from_coords_duckdb(conn, "csv_data", x_col, y_col) or 4326

            # Transform bounds to detected CRS if not WGS84
            if detected_crs_val != 4326:
                from pyproj import Transformer
                transformer = Transformer.from_crs("EPSG:4326", f"EPSG:{detected_crs_val}", always_xy=True)
                bminx, bminy = transformer.transform(bminx, bminy)
                bmaxx, bmaxy = transformer.transform(bmaxx, bmaxy)

            sample_df = conn.execute(f"""
                SELECT "{x_col}", "{y_col}" FROM csv_data
                WHERE "{x_col}" IS NOT NULL AND "{y_col}" IS NOT NULL
                  AND "{x_col}" BETWEEN {bminx} AND {bmaxx}
                  AND "{y_col}" BETWEEN {bminy} AND {bmaxy}
                USING SAMPLE 1000
            """).fetchdf()
            if len(sample_df) > 0:
                geometry = gpd.points_from_xy(sample_df[x_col], sample_df[y_col])
                last_gdf = gpd.GeoDataFrame(sample_df, geometry=geometry, crs=f"EPSG:{detected_crs_val}")
        except Exception as e:
            print(f"[DuckDB] Sample GeoDataFrame creation error: {e}")

    elif res_geom['columns'] and res_geom['method'] == 'linestring_coords':
        print(f"[DuckDB] Geometry detected: {res_geom['columns']} ({res_geom['method']})")

        x_start, y_start, x_end, y_end = res_geom['columns']
        geo_summary = process_geometry_duckdb_linestrings(conn, "csv_data", x_start, y_start, x_end, y_end, gdf_metro,
                                                           wgs84_bounds=wgs84_bounds, metric_crs=metric_crs)
        base_summary["Type de fichier"] = "CSV with Geometry (DuckDB Spatial)"

        try:
            sample_wkt = conn.execute("""
                SELECT ST_AsText(geom) as wkt FROM geo_processed
                USING SAMPLE 1000
            """).fetchdf()
            if len(sample_wkt) > 0:
                from shapely import wkt as shapely_wkt
                geometries = [shapely_wkt.loads(w) for w in sample_wkt['wkt'] if w]
                last_gdf = gpd.GeoDataFrame(geometry=geometries, crs="EPSG:4326")
        except Exception as e:
            print(f"[DuckDB] Sample GeoDataFrame creation error: {e}")

    elif res_geom['columns'] and res_geom['method'] in ['from_wkt', 'geojson', 'geopoint']:
        print(f"[DuckDB] Geometry detected: {res_geom['columns']} ({res_geom['method']})")
        df = conn.execute("SELECT * FROM csv_data").fetchdf()
        gdf = create_geodataframe_from_result(df, res_geom)
        gdf_proj, geo_metrics = process_geodataframe(gdf, gdf_metro)

        if geo_metrics:
            geo_summary = geo_metrics
            base_summary["Type de fichier"] = "CSV with Geometry (DuckDB)"
            last_gdf = gdf_proj
    
    # --- DateTime analysis ---
    datetime_analysis = analyze_datetime_columns(conn, "csv_data", raw_filepath=filepath)
    if datetime_analysis:
        base_summary["Analyse temporelle"] = datetime_analysis
        
    conn.close()

    total_time = time.time() - start_time
    print(f"[DuckDB] Total inspection time: {total_time:.2f}s")
    
    granularite = detect_granularite(
        ", ".join(k["col"] for k in geo_keys) if geo_keys else "None",
        geo_summary)
    
    summary_rows.append({
        **base_summary,
        **geo_summary,
        "Granularité": granularite,
    })
    print(f"\n{filepath} done\n")

# ============================================================================
# EXCEL HELPERS
# ============================================================================

def find_best_data_sheet(wb, sheet_names):
    """Find the sheet most likely to contain data."""
    if len(sheet_names) == 1:
        return sheet_names[0]
        
    metadata_patterns = ['readme', 'lisez', 'info', 'indic', 'description', 'metadata', 'note', 'about', 'legend', 'source', 'rapport']
    data_patterns = ['data', 'donnee', 'donnée', 'tableau', 'main', 'result', 'mesure', 'values', 'export']

    best_sheet = sheet_names[0]
    best_score = -1

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        score = 0
        name_lower = sheet_name.lower()

        if any(p in name_lower for p in metadata_patterns):
            score -= 10
        if any(p in name_lower for p in data_patterns):
            score += 10

        # score += min((ws.max_row or 0) / 100, 50)
        # score += min(ws.max_column or 0, 20)
        score += min((ws.max_row or 0) / 10, 100)

        if score > best_score:
            best_score = score
            best_sheet = sheet_name

    return best_sheet
    
def find_header_row(ws, max_rows_to_check=30):
    """Detect header row from worksheet."""
    rows_data = list(ws.iter_rows(max_row=max_rows_to_check, values_only=True))
    if not rows_data:
        return 0

    best_header_row = 0
    best_score = 0

    for row_idx, row in enumerate(rows_data[:-1]):
        non_empty = sum(1 for c in row if c is not None and str(c).strip())
        if non_empty < 2:
            continue

        string_cells = sum(
            1 for c in row
            if c is not None and isinstance(c, str)
            and not str(c).replace('.', '').replace('-', '').replace(',', '').isdigit()
        )

        next_row = rows_data[row_idx + 1]
        next_non_empty = sum(1 for c in next_row if c is not None and str(c).strip())

        score = non_empty + string_cells * 0.5 + (next_non_empty * 0.3 if next_non_empty > 0 else 0)
        if non_empty >= 3 and next_non_empty >= non_empty * 0.5:
            score += 5

        if score > best_score:
            best_score = score
            best_header_row = row_idx

    return best_header_row


def get_geo_columns(df):
    """Identify geometry columns and return structured info."""
    lat_pattern = r'\b(lat|latitude)\b'
    lon_pattern = r'\b(lon|long|lng|longitude)\b'
    x_pattern = r'(^x|[^a-zA-Z0-9]x$)'
    y_pattern = r'(^y|[^a-zA-Z0-9]y$)'
    geom_pattern = r'\b(geometry|geom|shape|point|polygon)\b'
    addr_pattern = r'adresse'
    insee_pattern = r'(dep|reg|insee|com)'

    result = {
        'columns': [],
        'type': None,
        'method': None,
        'geo_keys': [],
        'geotrans': 'Aucune géométrie',
    }

    # 1. Check for GeoJSON/WKT geometry columns
    for col in df.columns:
        col_lc = col.lower()
        if not re.search(geom_pattern, col_lc):
            continue

        sample = df[col].dropna().iloc[0] if not df[col].dropna().empty else None
        if sample is None:
            continue

        if 'point' in col_lc and 'geo' in col_lc and isinstance(sample, str) and ',' in sample:
            return {**result, 'type': 'Point', 'method': 'geopoint', 'columns': [col], 'geotrans': "Présence géométrie"}

        try:
            val = json.loads(sample) if isinstance(sample, str) else sample
            if isinstance(val, dict) and 'type' in val and 'coordinates' in val:
                return {**result, 'type': val.get('type', 'Unknown'), 'method': 'geojson', 'columns': [col], 'geotrans': "Présence géométrie"}
        except:
            pass

        if isinstance(sample, str):
            sample_upper = sample.upper()
            for geom_type in ['POINT', 'LINESTRING', 'POLYGON']:
                if geom_type in sample_upper:
                    return {**result, 'type': geom_type.title(), 'method': 'from_wkt', 'columns': [col], 'geotrans': "Présence géométrie"}

    # 2. Check for Lat/Lon pairs
    lat_col = lon_col = None
    for col in df.columns:
        col_lc = col.lower()
        if re.search(lat_pattern, col_lc) and lat_col is None:
            lat_col = col
        if re.search(lon_pattern, col_lc) and lon_col is None:
            lon_col = col

    if lat_col and lon_col:
        return {**result, 'type': 'Point', 'method': 'points_from_xy', 'columns': [lon_col, lat_col], 'geotrans': "Présence géométrie séparée (x,y)"}

    # 3. Check for X/Y pairs
    x_cols = [col for col in df.columns if re.search(x_pattern, col.lower())]
    y_cols = [col for col in df.columns if re.search(y_pattern, col.lower())]

    if x_cols and y_cols:
        for x_col in x_cols:
            for y_col in y_cols:
                x_suffix = x_col.lower().replace('x', '')
                y_suffix = y_col.lower().replace('y', '')

                if x_suffix == y_suffix or (not x_suffix and not y_suffix):
                    if any(k in x_col.lower() for k in ['d', 'debut', 'start']):
                        for x_end in x_cols:
                            if x_end != x_col and any(k in x_end.lower() for k in ['f', 'fin', 'end']):
                                for y_end in y_cols:
                                    if y_end != y_col and any(k in y_end.lower() for k in ['f', 'fin', 'end']):
                                        return {**result, 'type': 'LineString', 'method': 'linestring_coords', 
                                                'columns': [x_col, y_col, x_end, y_end], 'geotrans': "Présence géométrie multiples (x1,y1), (x2,y2)"}

                    return {**result, 'type': 'Point', 'method': 'points_from_xy', 'columns': [x_col, y_col], 'geotrans': "Présence géométrie séparée (x,y)"}

    # 4. Check for address columns
    for col in df.columns:
        if re.search(addr_pattern, col.lower()):
            return {**result, 'type': 'Address', 'method': 'geocoding_required', 'columns': [col], 'geotrans': "Géocodage de l'adresse"}

    # 5. Collect INSEE/geographic keys
    for col in df.columns:
        if re.search(insee_pattern, col.lower()):
            result['geo_keys'].append(col)

    if result['geo_keys']:
        result['type'] = 'Administrative'
        result['method'] = 'join_required'
        result['columns'] = result['geo_keys']
        result['geotrans'] = "Jointure spatiale à l'aide de clés géographiques"

    return result

    
def fix_insee_codes(df):
    """Fix INSEE codes that were read as integers."""
    fixed_columns = []
    pattern = r'(dep|reg|insee|com|code)'

    for col in df.columns:
        if not re.search(pattern, col.lower()):
            continue
        if not pd.api.types.is_numeric_dtype(df[col]):
            continue
        if df[col].notna().sum() == 0:
            continue

        avg_len = df[col].astype(str).str.len().mean()

        if 3.5 <= avg_len < 4.5:
            df[col] = df[col].apply(lambda x: str(int(x)).zfill(5) if pd.notna(x) else x)
            fixed_columns.append((col, 5, "code_insee_commune"))
        elif 1.0 <= avg_len < 1.5:
            df[col] = df[col].apply(lambda x: str(int(x)).zfill(2) if pd.notna(x) else x)
            fixed_columns.append((col, 2, "code_departement"))

    return df, fixed_columns


def create_geodataframe_from_result(df, res_geom):
    """Create GeoDataFrame from detection result."""
    method = res_geom['method']
    cols = res_geom['columns']

    if method == 'points_from_xy':
        x_col, y_col = cols
        geometry = gpd.points_from_xy(df[x_col], df[y_col])
        return gpd.GeoDataFrame(df, geometry=geometry)

    elif method == 'geojson':
        def parse_geojson(val):
            if pd.isna(val):
                return None
            try:
                geom_dict = json.loads(val) if isinstance(val, str) else val
                return shape(geom_dict)
            except:
                return None

        df = df.copy()
        df['geometry'] = df[cols[0]].apply(parse_geojson)
        return gpd.GeoDataFrame(df, geometry='geometry')

    elif method == 'geopoint':
        def parse_geopoint(val):
            if pd.isna(val) or not val:
                return None
            try:
                parts = str(val).split(',')
                if len(parts) == 2:
                    return Point(float(parts[1].strip()), float(parts[0].strip()))
            except:
                return None

        df = df.copy()
        df['geometry'] = df[cols[0]].apply(parse_geopoint)
        return gpd.GeoDataFrame(df, geometry='geometry')

    elif method == 'from_wkt':
        def parse_wkt(val):
            if pd.isna(val):
                return None
            try:
                return wkt.loads(str(val))
            except:
                return None

        df = df.copy()
        df['geometry'] = df[cols[0]].apply(parse_wkt)
        return gpd.GeoDataFrame(df, geometry='geometry')

    elif method == 'linestring_coords':
        x_start, y_start, x_end, y_end = cols

        def create_linestring(row):
            try:
                def to_float(val):
                    if pd.isna(val):
                        return None
                    if isinstance(val, str):
                        val = val.replace(',', '.')
                    return float(val)

                coords = [to_float(row[c]) for c in [x_start, y_start, x_end, y_end]]
                if any(v is None for v in coords):
                    return None
                return LineString([(coords[0], coords[1]), (coords[2], coords[3])])
            except:
                return None

        df = df.copy()
        df['geometry'] = df.apply(create_linestring, axis=1)
        return gpd.GeoDataFrame(df, geometry='geometry')

    else:
        raise ValueError(f"Unknown method: {method}")

        
def detect_granularite(geo_keys, geo_summary):
    """Déduit la granularité spatiale des données à partir des clés géo et des métriques."""
    
    granularites = []

    geom_types = geo_summary.get("Types de géométrie", "N/A")
    if geom_types and geom_types != "N/A":
        geom_lc = geom_types.lower()
        if "point" in geom_lc:
            granularites.append("Ponctuelle (géométrie)")
        elif "linestring" in geom_lc or "line" in geom_lc:
            granularites.append("Linéaire (géométrie)")
        elif "polygon" in geom_lc:
            granularites.append("Surfacique (géométrie)")

    if geo_keys and geo_keys != "Aucune":
        keys_lc = geo_keys.lower()
        if any(x in keys_lc for x in ["insee", "code_insee", "commune", "postal"]):
            granularites.append("Commune / INSEE")
        elif any(x in keys_lc for x in ["departement", "département", "code_departement"]):
            granularites.append("Département")
        elif any(x in keys_lc for x in ["region", "région", "code_region"]):
            granularites.append("Région")

    return " + ".join(granularites) if granularites else "Inconnue"
    
# ============================================================================
# DATETIME ANALYSIS
# ============================================================================

# Semantic keywords for interval pair detection
_INTERVAL_KEYWORDS = {
    "start": ["debut", "début", "start", "from", "ouverture", "open",
              "début_periode", "date_deb", "datedeb", "date_debut", "date_début", "datedébut", "datedéb"
              "date_from", "valid_from", "debut_validite"],
    "end":   ["fin", "end", "to", "fermeture", "close",
              "fin_periode", "date_fin", "datefin",
              "date_to", "valid_to", "fin_validite"],
}


def detect_date_columns_duckdb(conn, table_name):
    """
    Detect date columns in a DuckDB table by checking column types and values.
    Returns a list of column names that appear to contain dates.
    """
    schema = conn.execute(f"DESCRIBE {table_name}").fetchdf()
    date_cols = []
    for _, row in schema.iterrows():
        col = row["column_name"]
        t = row["column_type"].upper()
        if any(k in t for k in ("DATE", "TIMESTAMP", "TIME")):
            date_cols.append(col)
            continue
        # Also detect VARCHAR columns that look like dates
        if "VARCHAR" in t or "TEXT" in t:
            try:
                sample = conn.execute(f"""
                    SELECT CAST("{col}" AS VARCHAR) FROM {table_name}
                    WHERE "{col}" IS NOT NULL LIMIT 1
                """).fetchone()
                if sample and sample[0]:
                    s = str(sample[0]).strip()
                    if re.match(r'^\d{2}/\d{2}/\d{4}', s) or \
                       re.match(r'^\d{4}[-/]\d{2}[-/]\d{2}', s):
                        date_cols.append(col)
            except Exception:
                continue
    return date_cols


def classify_date_columns(date_cols):
    """
    Classify date columns into:
      - interval_pairs: [(col_start, col_end), ...]   → occupancy curve
      - singles:        [col, ...]                    → histogram

    Pair detection is purely semantic (column name keywords).
    Unpaired columns go to singles.
    """
    paired = set()
    interval_pairs = []

    cols_lower = {c: c.lower().replace(" ", "_") for c in date_cols}

    for col_s in date_cols:
        if col_s in paired:
            continue
        lc_s = cols_lower[col_s]
        if not any(kw in lc_s for kw in _INTERVAL_KEYWORDS["start"]):
            continue

        # Try to find a matching end column
        for col_e in date_cols:
            if col_e in paired or col_e == col_s:
                continue
            lc_e = cols_lower[col_e]
            if not any(kw in lc_e for kw in _INTERVAL_KEYWORDS["end"]):
                continue

            # Bonus: check they share a common root (e.g. "date_deb" / "date_fin")
            # by stripping the start/end keyword and comparing stems
            stem_s = lc_s
            stem_e = lc_e
            for kw in _INTERVAL_KEYWORDS["start"]:
                stem_s = stem_s.replace(kw, "")
            for kw in _INTERVAL_KEYWORDS["end"]:
                stem_e = stem_e.replace(kw, "")
            stem_s = stem_s.strip("_")
            stem_e = stem_e.strip("_")

            if stem_s == stem_e or len(stem_s) == 0 or len(stem_e) == 0:
                interval_pairs.append((col_s, col_e))
                paired.add(col_s)
                paired.add(col_e)
                break  # one match per start column

    singles = [c for c in date_cols if c not in paired]
    return interval_pairs, singles


def compute_occupancy_curve(conn, table_name, col_start, col_end, n_buckets=100, raw_formats=None):
    """
    Compute an occupancy curve for an interval pair (col_start, col_end).
    """
    raw_formats = raw_formats or {}

    def ts(col):
        fmt = raw_formats.get(col)
        if fmt is None:
            # Explicitly set to None = already correctly typed by DuckDB
            return f'CAST("{col}" AS TIMESTAMP)'
        if fmt:
            return f"TRY_STRPTIME(CAST(\"{col}\" AS VARCHAR), '{fmt}')"
        # Key absent = unknown, fallback to CAST
        return f'CAST("{col}" AS TIMESTAMP)'

    # --- Global temporal extent ---
    print(f"[DateTime] raw_formats={raw_formats}, ts(col_start)={ts(col_start)}, ts(col_end)={ts(col_end)}")
    # Verify TRY_STRPTIME works on a sample
    sample_check = conn.execute(f"""
        SELECT {ts(col_start)}, {ts(col_end)}
        FROM {table_name}
        WHERE "{col_start}" IS NOT NULL AND "{col_end}" IS NOT NULL
        LIMIT 3
    """).fetchall()
    print(f"[DateTime] sample parsed values: {sample_check}")
    
    extent = conn.execute(f"""
        SELECT
            MIN({ts(col_start)}) AS t_min,
            MAX({ts(col_end)})   AS t_max
        FROM {table_name}
        WHERE "{col_start}" IS NOT NULL
          AND "{col_end}"   IS NOT NULL
          AND YEAR({ts(col_start)}) >= 1800
          AND YEAR({ts(col_end)})   >= 1800
    """).fetchone()

    if extent is None or extent[0] is None or extent[1] is None:
        return None

    t_min, t_max = extent
    print(f"[DateTime] extent: t_min={t_min}, t_max={t_max}")
    coverage_days = (t_max - t_min).days if hasattr(t_max - t_min, "days") else 0

    # --- Choose time unit for the axis ---
    if coverage_days > 365:
        unit = "year"
    elif coverage_days > 31:
        unit = "month"
    else:
        unit = "day"

    # --- Median interval duration ---
    med_result = conn.execute(f"""
        SELECT MEDIAN(
            DATEDIFF('day', {ts(col_start)}, {ts(col_end)})
        )
        FROM {table_name}
        WHERE "{col_start}" IS NOT NULL
          AND "{col_end}"   IS NOT NULL
          AND YEAR({ts(col_start)}) >= 1800
          AND YEAR({ts(col_end)})   >= 1800
          AND {ts(col_end)} >= {ts(col_start)}
    """).fetchone()

    median_duration_days = float(med_result[0]) if med_result and med_result[0] is not None else 0

    # Intervalles <= 1 jour = mesures ponctuelles → histogramme simple
    if median_duration_days <= 1:
        result = compute_single_date_histogram(conn, table_name, col_start,
                                               n_buckets=n_buckets, raw_formats=raw_formats)
        if result:
            # Keep col_start/col_end so the JS renderer knows the pair origin
            result["col_start"] = col_start
            result["col_end"] = col_end
        return result

    # --- Build one bucket per period and count intersecting intervals ---
    unit_interval = {'year': 'YEAR', 'month': 'MONTH', 'day': 'DAY'}[unit]
    t_min_iso = t_min.isoformat()[:10]
    t_max_iso = t_max.isoformat()[:10]
    if unit == 'year':
        t_min_trunc = f"{t_min_iso[:4]}-01-01"
        n_periods   = int(t_max_iso[:4]) - int(t_min_iso[:4]) + 2
    elif unit == 'month':
        t_min_trunc = f"{t_min_iso[:7]}-01"
        import datetime as _dt_occ
        _d1 = _dt_occ.date.fromisoformat(t_min_trunc)
        _d2 = _dt_occ.date.fromisoformat(f"{t_max_iso[:7]}-01")
        n_periods = (_d2.year - _d1.year) * 12 + (_d2.month - _d1.month) + 2
    else:
        t_min_trunc = t_min_iso
        n_periods   = coverage_days + 2

    conn.execute(f"""
        CREATE OR REPLACE TEMP TABLE _time_axis AS
        SELECT TIMESTAMP '{t_min_trunc}' +
               (INTERVAL 1 {unit_interval} * CAST(generate_series AS BIGINT)) AS bucket_start
        FROM generate_series(0, {n_periods})
    """)

    conn.execute(f"""
        CREATE OR REPLACE TEMP TABLE _occupancy AS
        SELECT
            a.bucket_start,
            COUNT(d."{col_start}") AS active_count
        FROM _time_axis a
        LEFT JOIN {table_name} d
            ON {ts(col_start)} < a.bucket_start + INTERVAL 1 {unit_interval}
           AND {ts(col_end)}   >= a.bucket_start
           AND d."{col_start}" IS NOT NULL
           AND d."{col_end}"   IS NOT NULL
           AND YEAR({ts(col_start)}) >= 1800
           AND YEAR({ts(col_end)})   >= 1800
        GROUP BY a.bucket_start
        ORDER BY a.bucket_start
    """)

    rows = conn.execute("SELECT bucket_start, active_count FROM _occupancy").fetchall()

    conn.execute("DROP TABLE IF EXISTS _time_axis")
    conn.execute("DROP TABLE IF EXISTS _occupancy")

    buckets = [
        {"t": row[0].isoformat()[:10] if hasattr(row[0], "isoformat") else str(row[0])[:10],
         "count": int(row[1])}
        for row in rows
    ]

    return {
        "col_start": col_start,
        "col_end":   col_end,
        "global_min": t_min.isoformat(),
        "global_max": t_max.isoformat(),
        "coverage_days": coverage_days,
        "unit": unit,
        "median_duration_days": round(median_duration_days, 1),
        "chart_type": "occupancy",
        "buckets": buckets,
    }
    
def compute_single_date_histogram(conn, table_name, col, n_buckets=50, raw_formats=None):
    """
    Compute a histogram (count per time unit) for a single date column.
    Unit is adaptive: year / month / day depending on coverage.
    """
    raw_formats = raw_formats or {}

    def ts(col):
        fmt = raw_formats.get(col)
        if fmt is None:
            # Explicitly set to None = already correctly typed by DuckDB
            return f'CAST("{col}" AS TIMESTAMP)'
        if fmt:
            return f"TRY_STRPTIME(CAST(\"{col}\" AS VARCHAR), '{fmt}')"
        # Key absent = unknown, fallback to CAST
        return f'CAST("{col}" AS TIMESTAMP)'

    extent = conn.execute(f"""
        SELECT
            MIN({ts(col)}) AS t_min,
            MAX({ts(col)}) AS t_max
        FROM {table_name}
        WHERE "{col}" IS NOT NULL
        AND YEAR({ts(col)}) >= 1800
    """).fetchone()

    if extent is None or extent[0] is None or extent[1] is None:
        return None

    t_min, t_max = extent
    print(f"[DateTime] extent: t_min={t_min}, t_max={t_max}")
    
    coverage_days = (t_max - t_min).days if hasattr(t_max - t_min, "days") else 0

    if coverage_days > 365 * 2:
        trunc_unit, unit = "year", "year"
    elif coverage_days > 60:
        trunc_unit, unit = "month", "month"
    else:
        trunc_unit, unit = "day", "day"

    rows = conn.execute(f"""
        SELECT
            DATE_TRUNC('{trunc_unit}', {ts(col)}) AS period,
            COUNT(*) AS cnt
        FROM {table_name}
        WHERE "{col}" IS NOT NULL
        AND YEAR({ts(col)}) >= 1800
        GROUP BY period
        ORDER BY period
    """).fetchall()

    buckets = [
        {"t": row[0].isoformat() if hasattr(row[0], "isoformat") else str(row[0]),
         "count": int(row[1])}
        for row in rows
    ]

    return {
        "col": col,
        "global_min": t_min.isoformat(),
        "global_max": t_max.isoformat(),
        "coverage_days": coverage_days,
        "unit": unit,
        "chart_type": "histogram",
        "buckets": buckets,
    }


def analyze_datetime_columns(conn, table_name, raw_filepath=None):
    """
    Full datetime analysis for a table already loaded in DuckDB.
    Returns a dict with:
      - date_columns: list of all detected date column names
      - interval_pairs: list of occupancy curve results
      - singles: list of histogram results
    """
    date_cols = detect_date_columns_duckdb(conn, table_name)
    if not date_cols:
        return None

    # Pre-detect raw string formats from file before DuckDB parsing corrupts them
    raw_formats = {}
    if raw_filepath:
        try:
            # Detect separator from first line to avoid misparse
            with open(raw_filepath, 'rb') as _f:
                _first = _f.read(2048).decode('utf-8', errors='replace')
            _sep = ';' if _first.count(';') > _first.count(',') else ','
            df_raw = pd.read_csv(raw_filepath, nrows=5, dtype=str,
                                 sep=_sep, encoding='utf-8-sig', encoding_errors='replace')
            
            # Get DuckDB column types to know which are already typed
            schema = conn.execute(f"DESCRIBE {table_name}").fetchdf()
            duckdb_types = dict(zip(schema['column_name'], schema['column_type']))

            for col in date_cols:
                col_type = duckdb_types.get(col, '').upper()
                already_typed = any(k in col_type for k in ('DATE', 'TIMESTAMP'))

                if already_typed:
                    # DuckDB parsed it correctly — no strptime needed
                    raw_formats[col] = None
                    continue

                if col in df_raw.columns:
                    vals = df_raw[col].dropna()
                    if vals.empty:
                        continue
                    s = vals.iloc[0].strip()
                    if re.match(r'^\d{2}/\d{2}/\d{4}', s):
                        raw_formats[col] = '%d/%m/%Y'
                    elif re.match(r'^\d{4}/\d{2}/\d{2} \d{2}:\d{2}', s):
                        raw_formats[col] = '%Y/%m/%d %H:%M:%S'
                    elif re.match(r'^\d{4}/\d{2}/\d{2}', s):
                        raw_formats[col] = '%Y/%m/%d'
                    elif re.match(r'^\d{2}-\d{2}-\d{4}', s):
                        raw_formats[col] = '%d-%m-%Y'
            print(f"[DateTime] df_raw columns: {df_raw.columns.tolist()}")
            print(f"[DateTime] date_cols from DuckDB: {date_cols}")
            print(f"[DateTime] raw_formats built: {raw_formats}")
            
        except Exception as e:
            print(f"[DateTime] Raw format detection failed: {e}")

    interval_pairs_meta, singles_meta = classify_date_columns(date_cols)

    # Store fine-grained data for subsequent filtering while connection is open
    _store_raw_temporal(conn, table_name, date_cols, raw_formats, interval_pairs_meta)

    interval_results = []
    for col_s, col_e in interval_pairs_meta:
        result = compute_occupancy_curve(conn, table_name, col_s, col_e, raw_formats=raw_formats)
        if result:
            interval_results.append(result)

    single_results = []
    for col in singles_meta:
        result = compute_single_date_histogram(conn, table_name, col, raw_formats=raw_formats)
        if result:
            single_results.append(result)

    return {
            "_datetime": True,          # ← marqueur pour le renderer HTML
            "date_columns": date_cols,
            "interval_pairs": interval_results,
            "singles": single_results,
            "all_columns": last_dt_raw.get("all_columns", []),
        }


def _store_raw_temporal(conn, table_name, date_cols, raw_formats, interval_pairs):
    """
    Store fine-grained (daily or monthly) temporal data globally so
    recompute_temporal_charts() can re-aggregate with filters later.
    """
    global last_dt_raw
    last_dt_raw = {
        "singles": {}, "intervals": {},
        "all_date_cols": date_cols,
        "raw_formats":   dict(raw_formats),   # stored for /temporal_plot reuse
        "all_columns":   [],
    }

    # Store every column name + type so the frontend can build the Y-axis selector
    try:
        schema_df = conn.execute(f"DESCRIBE {table_name}").fetchdf()
        last_dt_raw["all_columns"] = [
            {"name": row["column_name"], "type": row["column_type"]}
            for _, row in schema_df.iterrows()
        ]
    except Exception as _exc:
        print(f"[DateTime] all_columns store failed: {_exc}")

    def _te(col):
        fmt = raw_formats.get(col)
        if fmt is None:
            return f'CAST("{col}" AS TIMESTAMP)'
        if fmt:
            return f"TRY_STRPTIME(CAST(\"{col}\" AS VARCHAR), '{fmt}')"
        return f'CAST("{col}" AS TIMESTAMP)'

    for col in date_cols:
        te = _te(col)
        try:
            ext = conn.execute(f"""
                SELECT MIN({te}), MAX({te})
                FROM {table_name}
                WHERE "{col}" IS NOT NULL AND YEAR({te}) >= 1800
            """).fetchone()
            if not ext or ext[0] is None:
                continue
            t_min, t_max = ext
            cov_days = (t_max - t_min).days if hasattr(t_max - t_min, "days") else 0
            raw_gran = "month" if cov_days > 365 * 10 else "day"
            rows = conn.execute(f"""
                SELECT DATE_TRUNC('{raw_gran}', {te}) AS period, COUNT(*) AS cnt
                FROM {table_name}
                WHERE "{col}" IS NOT NULL AND YEAR({te}) >= 1800
                GROUP BY period ORDER BY period
            """).fetchall()
            last_dt_raw["singles"][col] = {
                "granularity": raw_gran,
                "global_min":  t_min.isoformat(),
                "global_max":  t_max.isoformat(),
                "data": [
                    (r[0].isoformat() if hasattr(r[0], "isoformat") else str(r[0]), int(r[1]))
                    for r in rows
                ],
            }
        except Exception as exc:
            print(f"[DateTime] _store_raw_temporal failed for '{col}': {exc}")

    for col_s, col_e in interval_pairs:
        te_s, te_e = _te(col_s), _te(col_e)
        try:
            ext = conn.execute(f"""
                SELECT MIN({te_s}), MAX({te_e})
                FROM {table_name}
                WHERE "{col_s}" IS NOT NULL AND "{col_e}" IS NOT NULL
                  AND YEAR({te_s}) >= 1800 AND YEAR({te_e}) >= 1800
            """).fetchone()
            if not ext or ext[0] is None:
                continue
            t_min, t_max = ext
            key = f"{col_s}||{col_e}"
            last_dt_raw["intervals"][key] = {
                "col_start":  col_s,
                "col_end":    col_e,
                "global_min": t_min.isoformat(),
                "global_max": t_max.isoformat(),
                # Store raw interval endpoints so recompute can do period-intersection counts
                "raw_intervals": [
                    (r[0].isoformat() if hasattr(r[0], "isoformat") else str(r[0]),
                     r[1].isoformat() if hasattr(r[1], "isoformat") else str(r[1]))
                    for r in conn.execute(f"""
                        SELECT {te_s}, {te_e}
                        FROM {table_name}
                        WHERE "{col_s}" IS NOT NULL AND "{col_e}" IS NOT NULL
                          AND YEAR({te_s}) >= 1800 AND YEAR({te_e}) >= 1800
                    """).fetchall()
                    if r[0] is not None and r[1] is not None
                ],
            }
        except Exception as exc:
            print(f"[DateTime] _store_raw_temporal failed for interval ({col_s},{col_e}): {exc}")


def recompute_temporal_charts(columns=None, date_start=None, date_end=None, granularity=None):
    """
    Re-compute temporal charts from cached raw data with optional filters.

    columns    : list of column names to include (None = all detected columns)
    date_start : ISO date string lower bound  (None = no lower bound)
    date_end   : ISO date string upper bound  (None = no upper bound)
    granularity: 'year'|'month'|'day'|None   (None = auto from effective coverage)

    Returns a dict in the same format as analyze_datetime_columns, or None if no
    cached data exists.
    """
    import datetime
    from collections import defaultdict

    if last_dt_raw is None:
        return None

    def _parse(s):
        if not s:
            return None
        try:
            return datetime.date.fromisoformat(str(s)[:10])
        except Exception:
            return None

    def _auto_gran(days):
        if days > 365 * 2:
            return "year"
        if days > 60:
            return "month"
        return "day"

    f_start = _parse(date_start)
    f_end   = _parse(date_end)

    def _eff_range(g_min_str, g_max_str):
        g_min = _parse(g_min_str)
        g_max = _parse(g_max_str)
        eff_s = max(g_min, f_start) if (f_start and g_min) else (f_start or g_min)
        eff_e = min(g_max, f_end)   if (f_end   and g_max) else (f_end   or g_max)
        return eff_s, eff_e

    # ── Singles (histograms) ─────────────────────────────────────────────────
    singles_out = []
    for col, raw in last_dt_raw.get("singles", {}).items():
        if columns and col not in columns:
            continue
        eff_s, eff_e = _eff_range(raw.get("global_min", ""), raw.get("global_max", ""))
        if eff_s and eff_e and eff_s > eff_e:
            continue
        cov = (eff_e - eff_s).days if (eff_s and eff_e) else 0
        tgt = granularity or _auto_gran(cov)

        acc = defaultdict(int)
        for date_str, count in raw.get("data", []):
            d = _parse(date_str)
            if d is None or (eff_s and d < eff_s) or (eff_e and d > eff_e):
                continue
            key = (f"{d.year}-01-01" if tgt == "year"
                   else f"{d.year}-{d.month:02d}-01" if tgt == "month"
                   else date_str[:10])
            acc[key] += count

        singles_out.append({
            "col":           col,
            "global_min":    eff_s.isoformat() if eff_s else raw.get("global_min", "")[:10],
            "global_max":    eff_e.isoformat() if eff_e else raw.get("global_max", "")[:10],
            "coverage_days": cov,
            "unit":          tgt,
            "chart_type":    "histogram",
            "buckets":       [{"t": k, "count": v} for k, v in sorted(acc.items())],
        })

    # ── Interval pairs (occupancy curves) ────────────────────────────────────
    intervals_out = []
    for key_str, raw in last_dt_raw.get("intervals", {}).items():
        col_s = raw.get("col_start", "")
        col_e = raw.get("col_end", "")
        if columns and col_s not in columns and col_e not in columns:
            continue
        eff_s, eff_e = _eff_range(raw.get("global_min", ""), raw.get("global_max", ""))
        if eff_s and eff_e and eff_s > eff_e:
            continue
        cov = (eff_e - eff_s).days if (eff_s and eff_e) else 0
        tgt = granularity or _auto_gran(cov)

        # Occupancy: count distinct intervals that intersect each period bucket
        acc = defaultdict(int)
        for start_str, end_str in raw.get("raw_intervals", []):
            iv_s = _parse(start_str)
            iv_e = _parse(end_str)
            if iv_s is None or iv_e is None:
                continue
            cur  = iv_s if not eff_s or iv_s >= eff_s else eff_s
            stop = iv_e if not eff_e or iv_e <= eff_e else eff_e
            if cur > stop:
                continue
            if tgt == 'year':
                cur_k  = datetime.date(cur.year, 1, 1)
                stop_k = datetime.date(stop.year, 1, 1)
                bound_s = datetime.date(eff_s.year, 1, 1) if eff_s else None
                bound_e = datetime.date(eff_e.year, 1, 1) if eff_e else None
                while cur_k <= stop_k:
                    if (not bound_s or cur_k >= bound_s) and (not bound_e or cur_k <= bound_e):
                        acc[f"{cur_k.year}-01-01"] += 1
                    cur_k = datetime.date(cur_k.year + 1, 1, 1)
            elif tgt == 'month':
                cur_k  = datetime.date(cur.year, cur.month, 1)
                stop_k = datetime.date(stop.year, stop.month, 1)
                bound_s = datetime.date(eff_s.year, eff_s.month, 1) if eff_s else None
                bound_e = datetime.date(eff_e.year, eff_e.month, 1) if eff_e else None
                while cur_k <= stop_k:
                    if (not bound_s or cur_k >= bound_s) and (not bound_e or cur_k <= bound_e):
                        acc[f"{cur_k.year}-{cur_k.month:02d}-01"] += 1
                    m = cur_k.month % 12 + 1
                    y = cur_k.year + (1 if cur_k.month == 12 else 0)
                    cur_k = datetime.date(y, m, 1)
            else:  # day
                cur_k = cur
                while cur_k <= stop:
                    if (not eff_s or cur_k >= eff_s) and (not eff_e or cur_k <= eff_e):
                        acc[cur_k.isoformat()] += 1
                    cur_k += datetime.timedelta(days=1)

        intervals_out.append({
            "col_start":           col_s,
            "col_end":             col_e,
            "global_min":          eff_s.isoformat() if eff_s else raw.get("global_min", "")[:10],
            "global_max":          eff_e.isoformat() if eff_e else raw.get("global_max", "")[:10],
            "coverage_days":       cov,
            "unit":                tgt,
            "median_duration_days": 0,
            "chart_type":          "occupancy",
            "buckets":             [{"t": k, "count": v} for k, v in sorted(acc.items())],
        })

    return {
        "_datetime":      True,
        "date_columns":   last_dt_raw.get("all_date_cols", []),
        "interval_pairs": intervals_out,
        "singles":        singles_out,
    }


def compute_column_occupancy(filepath, file_ext, date_col, end_date_col=None,
                              value_col=None, aggregation="presence",
                              date_start=None, date_end=None,
                              granularity=None, raw_formats=None, n_buckets=100):
    """
    Compute presence/occupancy or value aggregation of a column over time.

    X-axis mode (selected by end_date_col):
    - Histogram     (end_date_col=None)  : one bar per time bucket using date_col
    - Occupancy     (end_date_col set)   : count of active intervals at each bucket

    Y-axis mode (selected by aggregation):
    - "presence" : COUNT of non-null value_col rows (or all rows when value_col=None)
    - "mean"/"sum"/"min"/"max" : numeric aggregate of value_col per bucket
      (occupancy curve mode always uses "presence" regardless of this setting)

    Parameters
    ----------
    filepath      : path to the file kept in PREVIEW_DIR
    file_ext      : ".csv", ".xlsx", ".geojson", etc.
    date_col      : start (or single) date column
    end_date_col  : end date column → switches to occupancy curve mode
    value_col     : column to measure; None = count all rows
    aggregation   : "presence"|"mean"|"sum"|"min"|"max"
    date_start    : ISO date string lower bound  (None = no filter)
    date_end      : ISO date string upper bound  (None = no filter)
    granularity   : "year"|"month"|"day"|None   (None = auto from time span)
    raw_formats   : optional {col: strptime_format} hints from prior inspection
    n_buckets     : number of time-axis points for occupancy curve mode

    Returns dict with "buckets" list of {t, count}, or None on error.
    """
    import duckdb as _duckdb
    import datetime as _dt
    from collections import defaultdict

    raw_formats = raw_formats or {}
    ext = file_ext.lower()

    conn = _duckdb.connect(':memory:')
    try:
        conn.execute("INSTALL spatial; LOAD spatial;")
    except Exception:
        pass

    try:
        safe = filepath.replace("\\", "/")
        if ext in ('.csv', '.txt'):
            conn.execute(f"""
                CREATE TABLE _d AS
                SELECT * FROM read_csv_auto('{safe}',
                    ignore_errors=true, header=true, sample_size=-1)
            """)
        elif ext == '.xlsx':
            import pandas as _pd
            _df = _pd.read_excel(filepath)
            conn.execute("CREATE TABLE _d AS SELECT * FROM _df")
        elif ext in ('.geojson', '.json', '.shp', '.gpkg'):
            import geopandas as _gpd
            _gdf = _gpd.read_file(filepath)
            _df = _gdf.drop(columns=['geometry'], errors='ignore')
            conn.execute("CREATE TABLE _d AS SELECT * FROM _df")
        else:
            return None

        # ── Build timestamp expression for any column ─────────────────────
        schema_df = conn.execute("DESCRIBE _d").fetchdf()
        col_types  = dict(zip(schema_df['column_name'], schema_df['column_type']))

        def _te(col):
            ct = col_types.get(col, '').upper()
            if any(k in ct for k in ('DATE', 'TIMESTAMP', 'TIME')):
                return f'CAST("{col}" AS TIMESTAMP)'
            if col in raw_formats and raw_formats[col]:
                return f"TRY_STRPTIME(CAST(\"{col}\" AS VARCHAR), '{raw_formats[col]}')"
            sample = conn.execute(
                f'SELECT "{col}" FROM _d WHERE "{col}" IS NOT NULL LIMIT 1'
            ).fetchone()
            s = str(sample[0]).strip() if sample else ''
            if re.match(r'^\d{2}/\d{2}/\d{4}', s):
                return f"TRY_STRPTIME(CAST(\"{col}\" AS VARCHAR), '%d/%m/%Y')"
            if re.match(r'^\d{4}/\d{2}/\d{2}', s):
                return f"TRY_STRPTIME(CAST(\"{col}\" AS VARCHAR), '%Y/%m/%d')"
            return f'CAST("{col}" AS TIMESTAMP)'

        te_s = _te(date_col)
        te_e = _te(end_date_col) if end_date_col else None

        # ── Global temporal extent ────────────────────────────────────────
        if end_date_col:
            ext_row = conn.execute(f"""
                SELECT MIN({te_s}), MAX({te_e}) FROM _d
                WHERE "{date_col}" IS NOT NULL AND "{end_date_col}" IS NOT NULL
                  AND YEAR({te_s}) >= 1800 AND YEAR({te_e}) >= 1800
            """).fetchone()
        else:
            ext_row = conn.execute(f"""
                SELECT MIN({te_s}), MAX({te_s}) FROM _d
                WHERE "{date_col}" IS NOT NULL AND YEAR({te_s}) >= 1800
            """).fetchone()

        if not ext_row or ext_row[0] is None:
            return None
        t_min, t_max = ext_row

        # ── Auto granularity ──────────────────────────────────────────────
        if not granularity:
            g_s = _dt.date.fromisoformat(t_min.isoformat()[:10]) if hasattr(t_min, 'isoformat') else None
            g_e = _dt.date.fromisoformat(t_max.isoformat()[:10]) if hasattr(t_max, 'isoformat') else None
            e_s = _dt.date.fromisoformat(date_start) if date_start else g_s
            e_e = _dt.date.fromisoformat(date_end)   if date_end   else g_e
            cov = (e_e - e_s).days if (e_s and e_e) else 0
            if cov > 365 * 2:   granularity = 'year'
            elif cov > 60:      granularity = 'month'
            else:               granularity = 'day'

        trunc = {'year': 'year', 'month': 'month', 'day': 'day'}.get(granularity, 'month')
        # Optional value-column presence filter (shared by both modes)
        val_join_cond = f'AND d."{value_col}" IS NOT NULL' if value_col else ''
        val_where     = f'AND "{value_col}" IS NOT NULL'   if value_col else ''

        # ── Occupancy curve mode ──────────────────────────────────────────
        if end_date_col:
            unit_interval = {'year': 'YEAR', 'month': 'MONTH', 'day': 'DAY'}.get(granularity, 'MONTH')
            t_min_iso = t_min.isoformat()[:10]
            t_max_iso = t_max.isoformat()[:10]
            if granularity == 'year':
                t_min_trunc = f"{t_min_iso[:4]}-01-01"
                n_periods   = int(t_max_iso[:4]) - int(t_min_iso[:4]) + 2
            elif granularity == 'day':
                t_min_trunc = t_min_iso
                cov_days    = (t_max - t_min).days if hasattr(t_max - t_min, 'days') else 365
                n_periods   = cov_days + 2
            else:  # month (default)
                t_min_trunc = f"{t_min_iso[:7]}-01"
                import datetime as _dt2
                _d1 = _dt2.date.fromisoformat(t_min_trunc)
                _d2 = _dt2.date.fromisoformat(f"{t_max_iso[:7]}-01")
                n_periods = (_d2.year - _d1.year) * 12 + (_d2.month - _d1.month) + 2

            conn.execute(f"""
                CREATE OR REPLACE TEMP TABLE _tax AS
                SELECT TIMESTAMP '{t_min_trunc}' +
                       (INTERVAL 1 {unit_interval} * CAST(generate_series AS BIGINT)) AS bucket_start
                FROM generate_series(0, {n_periods})
            """)
            conn.execute(f"""
                CREATE OR REPLACE TEMP TABLE _occ AS
                SELECT a.bucket_start, COUNT(d."{date_col}") AS active_count
                FROM _tax a
                LEFT JOIN _d d
                    ON {te_s} < a.bucket_start + INTERVAL 1 {unit_interval}
                   AND {te_e} >= a.bucket_start
                   AND d."{date_col}" IS NOT NULL AND d."{end_date_col}" IS NOT NULL
                   AND YEAR({te_s}) >= 1800 AND YEAR({te_e}) >= 1800
                   {val_join_cond}
                GROUP BY a.bucket_start ORDER BY a.bucket_start
            """)

            range_conds = []
            if date_start: range_conds.append(f"bucket_start >= TIMESTAMP '{date_start} 00:00:00'")
            if date_end:   range_conds.append(f"bucket_start <= TIMESTAMP '{date_end} 23:59:59'")
            where_str = ('WHERE ' + ' AND '.join(range_conds)) if range_conds else ''
            occ_rows = conn.execute(f"SELECT bucket_start, active_count FROM _occ {where_str}").fetchall()
            conn.execute("DROP TABLE IF EXISTS _tax; DROP TABLE IF EXISTS _occ")

            buckets = [
                {"t": r[0].isoformat()[:10] if hasattr(r[0], 'isoformat') else str(r[0])[:10],
                 "count": int(r[1])}
                for r in occ_rows if r[0] is not None
            ]
            chart_type = "occupancy"

        # ── Histogram mode ────────────────────────────────────────────────
        else:
            where_parts = [f'"{date_col}" IS NOT NULL', f'YEAR({te_s}) >= 1800']
            if date_start: where_parts.append(f"{te_s} >= TIMESTAMP '{date_start} 00:00:00'")
            if date_end:   where_parts.append(f"{te_s} <= TIMESTAMP '{date_end} 23:59:59'")

            use_agg = aggregation in ('mean', 'sum', 'min', 'max') and value_col

            if use_agg:
                # Numeric aggregation: only include rows where value_col is parseable
                where_parts.append(f'TRY_CAST("{value_col}" AS DOUBLE) IS NOT NULL')
                func = {'mean': 'AVG', 'sum': 'SUM', 'min': 'MIN', 'max': 'MAX'}[aggregation]
                select_val = f'{func}(TRY_CAST("{value_col}" AS DOUBLE)) AS val'
            else:
                # Presence / count mode
                if val_where: where_parts.append(f'"{value_col}" IS NOT NULL')
                select_val = 'COUNT(*) AS val'

            where = ' AND '.join(where_parts)
            rows = conn.execute(f"""
                SELECT DATE_TRUNC('{trunc}', {te_s}) AS period, {select_val}
                FROM _d WHERE {where}
                GROUP BY period ORDER BY period
            """).fetchall()

            def _to_count(v):
                if v is None: return 0
                try: return round(float(v), 4)
                except: return 0

            buckets = [
                {"t": r[0].isoformat() if hasattr(r[0], 'isoformat') else str(r[0]),
                 "count": _to_count(r[1])}
                for r in rows if r[0] is not None
            ]
            chart_type = "histogram"

        return {
            "date_col":     date_col,
            "end_date_col": end_date_col,
            "value_col":    value_col,
            "aggregation":  aggregation,
            "granularity":  granularity,
            "chart_type":   chart_type,
            "global_min":   t_min.isoformat() if hasattr(t_min, 'isoformat') else str(t_min),
            "global_max":   t_max.isoformat() if hasattr(t_max, 'isoformat') else str(t_max),
            "buckets":      buckets,
        }

    except Exception as exc:
        print(f"[DateTime] compute_column_occupancy failed: {exc}")
        import traceback; traceback.print_exc()
        return None
    finally:
        conn.close()


def inspect_excel(filepath, gdf_metro, sample_size=5000, geo_key_patterns=None, wgs84_bounds=None, metric_crs=None):
    """
    Inspect Excel file using original optimized reading + DuckDB spatial for geometry.
    """
    global last_gdf

    print(f"[DuckDB] Inspecting Excel: {filepath}")
    start_time = time.time()

    meta = get_file_metadata(filepath)

    import openpyxl

    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    sheet_names = wb.sheetnames
    best_sheet = find_best_data_sheet(wb, sheet_names)
    ws = wb[best_sheet]

    total_rows = ws.max_row or 0
    total_cols = ws.max_column or 0
    header_row = find_header_row(ws, max_rows_to_check=30)
    wb.close()

    total_data_rows = total_rows - header_row - 1
    duplicate_pct = 0

    print(f"[DuckDB] Sheet scan done - using '{best_sheet}' ({total_data_rows} rows)")

    read_start = time.time()

    if total_data_rows <= sample_size:
        df = pd.read_excel(filepath, sheet_name=best_sheet, header=header_row, engine='calamine')
        res_geom = None
    else:
        print(f"[DuckDB] Large file ({total_data_rows} rows). Using smart sampling...")

        df_sample = pd.read_excel(filepath, sheet_name=best_sheet, header=header_row, nrows=100, engine='openpyxl')
        df_sample.columns = [str(c).replace('\r\n', ' ').replace('\n', ' ').strip() for c in df_sample.columns]
        res_geom = get_geo_columns(df_sample)

        if res_geom['columns'] and res_geom['method'] in ['points_from_xy', 'linestring_coords']:
            coord_df = pd.read_excel(filepath, sheet_name=best_sheet, header=header_row,
                                    usecols=res_geom['columns'], engine='calamine')

            coord_df['_idx'] = coord_df.index
            duplicate_pct = coord_df.duplicated(subset=res_geom['columns'], keep='first').sum() / len(coord_df) * 100
            coord_unique = coord_df.drop_duplicates(subset=res_geom['columns'], keep='first')

            if len(coord_unique) > sample_size:
                coord_unique = coord_unique.sample(n=sample_size, random_state=42)

            indices = sorted(coord_unique['_idx'].tolist())
            skip_rows = set(range(header_row + 1, total_rows)) - set([header_row + 1 + i for i in indices])

            df = pd.read_excel(filepath, sheet_name=best_sheet, header=header_row,
                              skiprows=list(skip_rows), engine='calamine')
        else:
            df = pd.read_excel(filepath, sheet_name=best_sheet, header=header_row,
                              nrows=sample_size, engine='calamine')

    read_time = time.time() - read_start
    print(f"[DuckDB] Excel read in {read_time:.2f}s")

    df.columns = [str(c).replace('\r\n', ' ').replace('\n', ' ').strip() for c in df.columns]
    df, _ = fix_insee_codes(df)

    if res_geom is None:
        res_geom = get_geo_columns(df)
        
    # ── Detect ambiguous date columns before DuckDB ingestion ────────────
    excel_bad_date_cols = []
    try:
        df_probe = pd.read_excel(filepath, sheet_name=best_sheet,
                                 header=header_row, nrows=3, dtype=str, engine='openpyxl')
        df_probe.columns = [str(c).replace('\r\n', ' ').replace('\n', ' ').strip()
                            for c in df_probe.columns]
        for col in df_probe.columns:
            vals = df_probe[col].dropna()
            if vals.empty:
                continue
            s = str(vals.iloc[0]).strip()
            if re.match(r'^\d{2}/\d{2}/\d{4}', s) or \
               re.match(r'^\d{2}-\d{2}-\d{4}', s):
                excel_bad_date_cols.append(col)
    except Exception as e:
        print(f"[DateTime] Excel probe read failed: {e}")
    # ─────────────────────────────────────────────────────────────────────
    
    conn = get_duckdb_connection()
    conn.register('excel_data', df)
    conn.execute("CREATE TABLE excel_tbl AS SELECT * FROM excel_data")
    

    geo_keys = detect_geo_join_keys_duckdb(conn, "excel_data", geo_key_patterns=geo_key_patterns)
    res_geom = get_geo_columns_duckdb(conn, "excel_data")

    geo_key_cols = [k["col"] for k in geo_keys]
    geo_key_completeness = completeness_score_duckdb_cols(conn, "excel_data", geo_key_cols)

    # Detect temporal/date columns (preliminary - full analysis done by analyze_datetime_columns)
    date_cols_detected = detect_date_columns_duckdb(conn, "excel_tbl")
    if date_cols_detected:
        print(f"[DuckDB] Date columns detected: {date_cols_detected}")

    sheet_info = f"Feuilles: {len(sheet_names)}, analysée: {best_sheet}"
    if header_row > 0:
        sheet_info += f", en-tête ligne {header_row + 1}"

    base_summary = {
        **meta,
        "Type de fichier": f"EXCEL (DuckDB Spatial) ({sheet_info})",
        "Nb lignes": total_data_rows,
        "Nb colonnes": total_cols,
        "Colonnes": {"_table": True, "data": build_columns_detail_duckdb(conn, "excel_tbl")},
        "Score de complétude global": completeness_score_duckdb(conn, "excel_tbl"),
        "Analyse temporelle": None,  # Will be populated by analyze_datetime_columns
        "Clés géographiques": format_geo_keys_table(geo_keys),
        "Géotransformation": res_geom['geotrans'],
        "Score de complétude des clés géographique": geo_key_completeness,
    }

    geo_summary = get_default_geo_summary()

    if res_geom['columns'] and res_geom['method'] == 'points_from_xy':
        print(f"[DuckDB] Geometry: {res_geom['columns']} ({res_geom['method']})")

        x_col, y_col = res_geom['columns']
        geo_summary = process_geometry_duckdb_points(conn, "excel_tbl", x_col, y_col, gdf_metro,
                                                      wgs84_bounds=wgs84_bounds, metric_crs=metric_crs)
        geo_summary['Part des geometries dupliquees (%)'] = round(duplicate_pct, 2)
        base_summary["Type de fichier"] = f"EXCEL with Geometry (DuckDB Spatial) ({sheet_info})"

        # Sample from geo_processed (already filtered, never reprojected)
        try:
            #### DEBUG 
            print(f"[DEBUG] geo_summary after points: {geo_summary.get('CRS')}, {geo_summary.get('Types de géométrie')}")
            try:
                tables = conn.execute("SHOW TABLES").fetchdf()
                print(f"[DEBUG] tables in conn: {tables['name'].tolist()}")
                count = conn.execute("SELECT COUNT(*) FROM geo_processed").fetchone()[0]
                print(f"[DEBUG] geo_processed row count: {count}")
            except : 
                print('error')
            #### DEBUG 
            sample_wkt = conn.execute("""
                SELECT ST_AsText(geom) as wkt FROM geo_processed
                USING SAMPLE 1000
            """).fetchdf()
            if len(sample_wkt) > 0:
                from shapely import wkt as shapely_wkt
                geometries = [shapely_wkt.loads(w) for w in sample_wkt['wkt'] if w]
                detected_crs_val = guess_crs_from_coords_duckdb(conn, "excel_tbl", x_col, y_col) or 4326
                last_gdf = gpd.GeoDataFrame(geometry=geometries, crs=f"EPSG:{detected_crs_val}")
        except Exception as e:
            print(f"[DuckDB] Sample GeoDataFrame creation error: {e}")

    elif res_geom['columns'] and res_geom['method'] == 'linestring_coords':
        print(f"[DuckDB] Geometry: {res_geom['columns']} ({res_geom['method']})")

        x_start, y_start, x_end, y_end = res_geom['columns']
        geo_summary = process_geometry_duckdb_linestrings(conn, "excel_tbl", x_start, y_start, x_end, y_end, gdf_metro,
                                                           wgs84_bounds=wgs84_bounds, metric_crs=metric_crs)
        geo_summary['Part des geometries dupliquees (%)'] = round(duplicate_pct, 2)
        base_summary["Type de fichier"] = f"EXCEL with Geometry (DuckDB Spatial) ({sheet_info})"

        try:
            sample_wkt = conn.execute("""
                SELECT ST_AsText(geom) as wkt FROM geo_processed
                USING SAMPLE 1000
            """).fetchdf()
            if len(sample_wkt) > 0:
                from shapely import wkt as shapely_wkt
                geometries = [shapely_wkt.loads(w) for w in sample_wkt['wkt'] if w]
                last_gdf = gpd.GeoDataFrame(geometry=geometries, crs="EPSG:4326")
        except Exception as e:
            print(f"[DuckDB] Sample GeoDataFrame creation error: {e}")

    elif res_geom['columns'] and res_geom['method'] in ['from_wkt', 'geojson', 'geopoint']:
        print(f"[DuckDB] Geometry: {res_geom['columns']} ({res_geom['method']})")
        gdf = create_geodataframe_from_result(df, res_geom)
        gdf_proj, geo_metrics = process_geodataframe(gdf, gdf_metro, compute_duplicates=False)

        if geo_metrics:
            geo_metrics['Part des geometries dupliquees (%)'] = round(duplicate_pct, 2)
            geo_summary = geo_metrics
            base_summary["Type de fichier"] = f"EXCEL with Geometry (DuckDB) ({sheet_info})"
            last_gdf = gdf_proj
            
    # --- DateTime analysis ---
    datetime_analysis = analyze_datetime_columns(conn, "excel_tbl", raw_filepath=filepath)

    if datetime_analysis:
        base_summary["Analyse temporelle"] = datetime_analysis
        
    conn.close()

    total_time = time.time() - start_time
    print(f"[DuckDB] Total inspection time: {total_time:.2f}s")

    granularite = detect_granularite(
            ", ".join(k["col"] for k in geo_keys) if geo_keys else "None",
            geo_summary)
    
    summary_rows.append({
        **base_summary,
        **geo_summary,
        "Granularité": granularite,
    })   
    print(f"\n{filepath} done\n")


def inspect_geospatial_duckdb(filepath, gdf_metro, geo_key_patterns=None, wgs84_bounds=None, metric_crs=None):
    """
    Inspect geospatial file using DuckDB spatial extension for both reading and processing.
    """
    global last_gdf
    os.environ["SHAPE_RESTORE_SHX"] = "YES"

    print(f"[DuckDB] Inspecting geospatial file: {filepath}")
    start_time = time.time()

    conn = get_duckdb_connection()

    try:        

        # ── Detect ambiguous date columns before DuckDB ingestion ────────────
        geo_bad_date_cols = []
        try:
            with open(filepath, 'rb') as _f:
                _first = _f.read(2048).decode('utf-8', errors='replace')
            _sep = ';' if _first.count(';') > _first.count(',') else ','
            df_probe = pd.read_csv(filepath, nrows=3, dtype=str,
                                   sep=_sep, encoding_errors='replace')
            for col in df_probe.columns:
                vals = df_probe[col].dropna()
                if vals.empty:
                    continue
                s = str(vals.iloc[0]).strip()
                if re.match(r'^\d{2}/\d{2}/\d{4}', s) or \
                   re.match(r'^\d{2}-\d{2}-\d{4}', s):
                    geo_bad_date_cols.append(col)
        except Exception as e:
            print(f"[DateTime] Geospatial probe read failed: {e}")
        # ─────────────────────────────────────────────────────────────────────
        
        conn.execute(f"""
            CREATE TABLE geo_data AS
            SELECT * FROM st_read('{filepath}')
        """)
        # Only re-read if ambiguous date cols detected AND it's a CSV-like format
        ext = os.path.splitext(filepath)[-1].lower()
        if geo_bad_date_cols and ext in ('.csv', '.txt'):
            print(f"[DateTime] Geospatial ambiguous date columns: {geo_bad_date_cols} — forcing VARCHAR")
            col_types = ", ".join([f'"{c}": "VARCHAR"' for c in geo_bad_date_cols])
            conn.execute("DROP TABLE IF EXISTS geo_data")
            try:
                conn.execute(f"""
                    CREATE TABLE geo_data AS
                    SELECT * FROM read_csv('{filepath}',
                        header=true,
                        auto_detect=true,
                        ignore_errors=true,
                        sample_size=100000,
                        types={{{col_types}}}
                    )
                """)
                print(f"[DateTime] Geospatial re-read done, forced VARCHAR for: {geo_bad_date_cols}")
            except Exception as e:
                print(f"[DateTime] Geospatial re-read failed: {e}")
            
        read_time = time.time() - start_time
        print(f"[DuckDB] Geospatial file read in {read_time:.2f}s")

        meta = get_file_metadata(filepath)

        row_count = conn.execute("SELECT COUNT(*) FROM geo_data").fetchone()[0]
        schema = conn.execute("DESCRIBE geo_data").fetchdf()
        col_count = len(schema)

        # Find geometry column
        geom_col = None
        geom_name_candidates = ['geom', 'geometry', 'wkb_geometry', 'shape', 'the_geom']
        col_names_lower = {col.lower(): col for col in schema['column_name'].tolist()}
        for candidate in geom_name_candidates:
            if candidate in col_names_lower:
                geom_col = col_names_lower[candidate]
                break

        if geom_col is None:
            print(f"[WARN] Aucune colonne géométrique trouvée. Colonnes disponibles: {schema['column_name'].tolist()}")

        geo_keys = detect_geo_join_keys_duckdb(conn, "geo_data", geo_key_patterns=geo_key_patterns)
        res_geom = get_geo_columns_duckdb(conn, "geo_data")

        geo_key_cols = [k["col"] for k in geo_keys]
        geo_key_completeness = completeness_score_duckdb_cols(conn, "geo_data", geo_key_cols)

        # Detect temporal/date columns (preliminary - full analysis done by analyze_datetime_columns)
        date_cols_detected = detect_date_columns_duckdb(conn, "geo_data")
        if date_cols_detected:
            print(f"[DuckDB] Date columns detected: {date_cols_detected}")

        completeness = completeness_score_duckdb(conn, "geo_data")
        columns_detail = build_columns_detail_duckdb(conn, "geo_data")

        if geom_col is not None:
            geo_summary = process_geometry_duckdb_native(conn, "geo_data", geom_col, gdf_metro,
                                                          wgs84_bounds=wgs84_bounds, metric_crs=metric_crs)
        else:
            geo_summary = get_default_geo_summary()

        # # Create sample GeoDataFrame for map display
        # try:
        #     sample_wkt = conn.execute("""
        #         SELECT ST_AsText(geom) as wkt FROM geo_processed
        #         USING SAMPLE 1000
        #     """).fetchdf()

        #     if len(sample_wkt) > 0:
        #         from shapely import wkt as shapely_wkt
        #         geometries = [shapely_wkt.loads(w) for w in sample_wkt['wkt'] if w]
        #         # Geometry is never reprojected — use detected source CRS
        #         import re as _re
        #         _crs_match = _re.search(r"EPSG:(\d+)", geo_summary.get("CRS", ""))
        #         detected_native_crs = int(_crs_match.group(1)) if _crs_match else (
        #             guess_crs_from_bounds_duckdb(conn, "geo_processed", "geom") or 4326
        #         )
        #         print(f"[DEBUG] detected_native_crs = {detected_native_crs}")
        #         last_gdf = gpd.GeoDataFrame(geometry=geometries, crs=f"EPSG:{detected_native_crs}").to_crs(epsg=4326)
                
        # except Exception as e:
        #     print(f"[DuckDB] Sample GeoDataFrame creation error: {e}")
        #     gdf = gpd.read_file(filepath)
        #     if len(gdf) > 1000:
        #         gdf = gdf.sample(n=1000, random_state=42)
        #     last_gdf = gdf
        
        # Create sample GeoDataFrame for map display

        try:
            sample_wkt = conn.execute("""
                SELECT ST_AsText(geom) as wkt FROM geo_processed
                USING SAMPLE 1000
            """).fetchdf()

            if len(sample_wkt) > 0:
                from shapely import wkt as shapely_wkt
                geometries = [shapely_wkt.loads(w) for w in sample_wkt['wkt'] if w]
                temp_gdf = gpd.GeoDataFrame(geometry=geometries)
                bounds = temp_gdf.total_bounds  # [minx, miny, maxx, maxy]
                median_x = (bounds[0] + bounds[2]) / 2
                median_y = (bounds[1] + bounds[3]) / 2
                print(f"[DEBUG] last_gdf median_x={median_x:.1f}, median_y={median_y:.1f}")
                # if -10 < median_x < 10 and 40 < median_y < 60:
                #     detected_native_crs = 4326
                # elif 100000 < median_x < 1300000 and 6000000 < median_y < 7400000:
                #     detected_native_crs = 2154
                # elif 0 < median_x < 800000 and 0 < median_y < 1300000:
                #     detected_native_crs = 27700
                # elif 0 < median_x < 1000000 and 5400000 < median_y < 7800000:
                #     detected_native_crs = 25832
                # elif -1000000 < median_x < 4000000 and 1500000 < median_y < 5500000:
                #     detected_native_crs = 3035
                # elif -700000 < median_x < 1100000 and 3900000 < median_y < 4900000:
                #     detected_native_crs = 2062
                # elif -2.2e6 < median_x < 2.2e6 and 3.0e6 < median_y < 2.2e7:
                #     detected_native_crs = 3857
                # else:
                #     detected_native_crs = 4326

                import fiona
                with fiona.open(filepath) as src:
                    file_crs = src.crs_wkt
                from pyproj import CRS
                detected_native_crs = CRS.from_wkt(file_crs).to_epsg() or 4326


                last_gdf = gpd.GeoDataFrame(geometry=geometries, crs=f"EPSG:{detected_native_crs}")
                
                print(f"[DuckDB] last_gdf CRS: {last_gdf.crs}, bounds: {last_gdf.total_bounds}")
        except Exception as e:
            print(f"[DuckDB] Sample GeoDataFrame creation error: {e}")
            last_gdf = None


        # --- DateTime analysis ---
        datetime_analysis = analyze_datetime_columns(conn, "geo_data", raw_filepath=filepath)
        conn.close()

        base_summary = {
            **meta,
            "Type de fichier": "Geospatial (DuckDB Spatial)",
            "Nb lignes": row_count,
            "Nb colonnes": col_count,
            "Colonnes": {"_table": True, "data": columns_detail},
            "Score de complétude global": completeness,
            "Analyse temporelle": None,  # Will be populated by analyze_datetime_columns
            "Clés géographiques": format_geo_keys_table(geo_keys),
            "Géotransformation": "Données géographiques",
            "Score de complétude des clés géographique": geo_key_completeness
        }

        if datetime_analysis:
            base_summary["Analyse temporelle"] = datetime_analysis
            
        granularite = detect_granularite(
            ", ".join(k["col"] for k in geo_keys) if geo_keys else "None",
            geo_summary)
        
        summary_rows.append({
            **base_summary,
            **geo_summary,
            "Granularité": granularite,
        })
        total_time = time.time() - start_time
        print(f"[DuckDB] Total inspection time: {total_time:.2f}s")
        print(f"\n{filepath} done\n")

    except Exception as e:
        conn.close()
        print(f"[DuckDB] Error: {e}")
        import traceback
        traceback.print_exc()
        raise


def guess_crs_from_bounds(gdf):
    """Guess CRS from GeoDataFrame using score-based detection."""
    if gdf.empty or gdf.geometry.is_empty.all():
        return None
    try:
        if gdf.geometry.iloc[0].geom_type == 'Point':
            xs, ys = gdf.geometry.x.values, gdf.geometry.y.values
        else:
            centroids = gdf.geometry.centroid
            xs, ys = centroids.x.values, centroids.y.values
        epsg, _conf, _alts = guess_crs_from_xy(xs, ys)
        return epsg
    except Exception:
        return None


def process_geodataframe(gdf, gdf_metro, compute_duplicates=True):
    """Process GeoDataFrame: reproject, validate, compute metrics."""
    try:
        if gdf.crs is None:
            epsg_crs = guess_crs_from_bounds(gdf)
            if epsg_crs:
                gdf = gdf.set_crs(epsg=epsg_crs)

        source_crs_str = gdf.crs.to_string() if gdf.crs else None

        if gdf.crs is not None and gdf.crs.to_string() != "EPSG:2154":
            gdf_proj = gdf.to_crs(epsg=2154)
        else:
            gdf_proj = gdf

        non_empty = (~gdf_proj.geometry.is_empty & gdf_proj.geometry.notna()).sum()
        valid_count = gdf_proj.geometry.dropna().apply(lambda g: g.is_valid).sum()
        total = len(gdf_proj)

        invalid_mask = ~gdf_proj.geometry.is_valid
        if invalid_mask.any():
            gdf_proj = gdf_proj.copy()
            gdf_proj.loc[invalid_mask, 'geometry'] = gdf_proj.loc[invalid_mask, 'geometry'].apply(make_valid)

        merged = gdf_proj.geometry.union_all()
        hull = gpd.GeoSeries([merged], crs=gdf_proj.crs).concave_hull(ratio=0.1, allow_holes=False).iloc[0]
        area_km2 = hull.area / 1e6 if hull and hull.area > 0 else 0
        density = len(gdf_proj) / area_km2 if area_km2 > 0 else 0

        remplissage = taux_de_remplissage(gdf_proj, gdf_metro)
        complexite = complexite_moyenne(gdf_proj)
        doublons = pourcentage_geometries_dupliquees(gdf_proj) if compute_duplicates else {'Part des geometries dupliquees (%)': 0}

        geo_metrics = {
        "Score de complétude géographique": {
            "_table": True,
            "data": [{
                "Présentes (%)": round(non_empty / total * 100, 1),
                "Valides (%)":   round(valid_count / total * 100, 1),
            }]
        },
            "CRS": f"Detected CRS - {source_crs_str} (transformed to {gdf_proj.crs.to_string()})" if source_crs_str else (gdf_proj.crs.to_string() if gdf_proj.crs else "Non défini"),
            "Types de géométrie": ", ".join(gdf_proj.geom_type.value_counts().index.tolist()),
            "Emprise estimée (km2)": round(area_km2, 2),
            "Densité (obj/km2)": round(density, 2),
            "Taux de remplissage géométrique (%)": remplissage['Taux de remplissage (%)'],
            "Complexité moyenne des géométries": complexite['Complexite moyenne (sommets)'],
            "Part des geometries dupliquees (%)": doublons['Geometries dupliquees (%)'],
            "Couverture territoriale (%)": remplissage['Couverture territoriale (%)'],
        }

        return gdf_proj, geo_metrics

    except Exception as e:
        print(f"Error processing GeoDataFrame: {e}")
        return None, None


def get_default_geo_summary():
    """Return default geo summary with N/A values."""
    return {
        "Score de complétude des clés géographique": {"_table": True, "data": [{"Score de complétude moyen (%)": "N/A", "Score de complétude std (%)": "N/A"}]},
        "CRS": "N/A",
        "Types de géométrie": "N/A",
        "Emprise estimée (km2)": None,
        "Densité (obj/km2)": None,
        "Taux de remplissage géométrique (%)": None,
        "Complexité moyenne des géométries": None,
        "Part des geometries dupliquees (%)": None,
        "Couverture territoriale (%)": None,
    }


# ============================================================================
# MAIN ENTRY POINT
# ============================================================================
def inspect_file(filepath, gdf_metro, geo_key_patterns=None, wgs84_bounds=None, metric_crs=None):
    """Main entry point - dispatch to appropriate DuckDB inspector."""
    ext = os.path.splitext(filepath)[-1].lower()

    if ext in ['.csv', '.txt']:
        inspect_csv_duckdb(filepath, gdf_metro, geo_key_patterns=geo_key_patterns, wgs84_bounds=wgs84_bounds, metric_crs=metric_crs)
    elif ext in ['.geojson', '.json', '.shp', '.gpkg']:
        inspect_geospatial_duckdb(filepath, gdf_metro, geo_key_patterns=geo_key_patterns, wgs84_bounds=wgs84_bounds, metric_crs=metric_crs)
    elif ext == ".xlsx":
        inspect_excel(filepath, gdf_metro, geo_key_patterns=geo_key_patterns, wgs84_bounds=wgs84_bounds, metric_crs=metric_crs)


# ============================================================================
# CLI
# ============================================================================
if __name__ == "__main__":
    import sys

    # Load reference data
    REFERENCE_PATH = os.path.join(os.path.dirname(__file__), "data", "regions.geojson")
    gdf_reference = gpd.read_file(REFERENCE_PATH).to_crs(epsg=2154)

    if len(sys.argv) < 2:
        print("Usage:")
        print("  python core.py <file_path>           # Inspect file")
        print()
        print("Examples:")
        print("  python core.py data/pesticides-2002-2022-v03-2024-vf.csv")
        sys.exit(1)

    else:
        inspect_file(sys.argv[1], gdf_reference, geo_key_patterns=None, wgs84_bounds=None, metric_crs=None)

        if summary_rows:
            print("\n--- Summary ---")
            for key, value in summary_rows[-1].items():
                if key != "Colonnes":
                    print(f"  {key}: {value}")